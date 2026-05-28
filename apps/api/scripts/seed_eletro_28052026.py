"""One-off: seed da categoria Eletro a partir de i.eletro.xlsx.

Importa produtos (aba "eletro") + cotação (aba "cotacao eletro") com
categoria='eletro'. A aba "resumo..." está vazia no arquivo — ignorada.
A aba "kit" do eletro é cópia do template do mala (sem dados úteis) —
ignorada (eletro não tem kit).

Idempotente: pula produtos/cotação se já houver registros eletro.

Estrutura do Excel (inspecionada 2026-05-28):
  eletro: header row 4 → A=fornecedor B=modelo_bling C=sku D=custo
          E=estoque F=consumo_diario G=memoria(maior_media) J=obs.
          Dados a partir da row 5 (12 SKUs preenchidos).
  cotacao eletro: fabricantes tywit(cols 2-4)/anbolife(5-7)/suplas(8-10);
          produtos na col 1 a partir da row 7; cada bloco = (capacidade,
          R$, USD).

Uso:
  docker compose -f docker-compose.yml -f docker-compose.prod.yml \
    exec -T api python -m scripts.seed_eletro_28052026
"""
from __future__ import annotations

import asyncio
import os
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import func, select

from app.db import session_scope
from app.models import (
    CotacaoFabricante,
    CotacaoProduto,
    CotacaoValor,
    ImportProduct,
)

# Path do Excel — override via env ELETRO_XLSX (ex.: /tmp/i.eletro.xlsx
# quando rodando dentro do container de prod).
XLSX_PATH = os.environ.get("ELETRO_XLSX", "/Users/admmarketing/Downloads/i.eletro.xlsx")

# (nome, primeira coluna do bloco capacidade/R$/USD)
_FABRICANTES = [("tywit", 2), ("anbolife", 5), ("suplas", 8)]


def _s(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _dec(v) -> Decimal | None:
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v).replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def _int(v) -> int | None:
    d = _dec(v)
    return int(d) if d is not None else None


async def _seed_products(session, wb) -> int:
    sh = wb["eletro"]
    added = 0
    for r in range(5, sh.max_row + 1):
        sku = _s(sh.cell(row=r, column=3).value)
        if not sku:
            continue
        session.add(ImportProduct(
            categoria="eletro",
            fornecedor=_s(sh.cell(row=r, column=1).value),
            modelo_bling=_s(sh.cell(row=r, column=2).value),
            sku=sku,
            custo_bling=_dec(sh.cell(row=r, column=4).value) or Decimal("0"),
            estoque_bling=_int(sh.cell(row=r, column=5).value),
            consumo_diario=_dec(sh.cell(row=r, column=6).value),
            maior_media_30d=_dec(sh.cell(row=r, column=7).value),
            obs=_s(sh.cell(row=r, column=10).value),
        ))
        added += 1
    return added


async def _seed_cotacao(session, wb) -> tuple[int, int, int]:
    sh = wb["cotacao eletro"]
    # Fabricantes.
    fab_by_name: dict[str, CotacaoFabricante] = {}
    for ordem, (nome, _col) in enumerate(_FABRICANTES):
        f = CotacaoFabricante(categoria="eletro", nome=nome, ordem=ordem)
        session.add(f)
        fab_by_name[nome] = f
    await session.flush()

    # Produtos (col 1, rows 7+).
    prods: list[tuple[int, CotacaoProduto]] = []
    ordem = 0
    for r in range(7, sh.max_row + 1):
        nome = _s(sh.cell(row=r, column=1).value)
        if not nome:
            continue
        p = CotacaoProduto(categoria="eletro", nome=nome, ordem=ordem)
        session.add(p)
        prods.append((r, p))
        ordem += 1
    await session.flush()

    # Valores: para cada produto × fabricante, lê o bloco (cap, R$, USD).
    n_val = 0
    for row_idx, prod in prods:
        for nome, col in _FABRICANTES:
            cap = _s(sh.cell(row=row_idx, column=col).value)
            vr = _dec(sh.cell(row=row_idx, column=col + 1).value)
            vu = _dec(sh.cell(row=row_idx, column=col + 2).value)
            if cap is None and vr is None and vu is None:
                continue
            session.add(CotacaoValor(
                fabricante_id=fab_by_name[nome].id,
                produto_id=prod.id,
                capacidade=cap,
                valor_real=vr,
                valor_usd=vu,
            ))
            n_val += 1
    return len(fab_by_name), len(prods), n_val


async def main() -> None:
    wb = load_workbook(XLSX_PATH, data_only=True)
    async with session_scope() as session:
        existing = (await session.execute(
            select(func.count()).select_from(ImportProduct)
            .where(ImportProduct.categoria == "eletro")
        )).scalar_one()
        if existing:
            print(f"⚠ já existem {existing} produtos eletro — pulando seed (idempotente).")
            return

        n_prod = await _seed_products(session, wb)
        n_fab, n_cot_prod, n_val = await _seed_cotacao(session, wb)
        await session.commit()
        print("✅ Seed eletro concluído")
        print(f"   produtos        : {n_prod}")
        print(f"   cotacao fab     : {n_fab}")
        print(f"   cotacao produtos: {n_cot_prod}")
        print(f"   cotacao valores : {n_val}")


if __name__ == "__main__":
    asyncio.run(main())
