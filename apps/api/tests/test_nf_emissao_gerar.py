"""Endpoint de geração da planilha de importação avulsa (Fase 3a).

`POST /api/nf-cadastro/faturamento/gerar-planilha` lê os itens do Bling
principal (bling_orders) já no davinci, aplica a regra do faturador da loja de
cada pedido e devolve o CSV no layout do relatório de vendas do Bling. Pedidos
sem faturador/itens saem em headers; 422 se nenhum gerou.
"""

from __future__ import annotations

import base64
import csv
import io
import json
import uuid
from collections.abc import Callable
from datetime import UTC, datetime
from decimal import Decimal

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    BlingOrder,
    ImportProduct,
    NfCatalogoMala,
    NfFaturador,
    Product,
    StoreInfo,
    User,
    UserRole,
    UserStatus,
)
from app.services import nf_upseller
from app.services.nf_relatorio import COLUNAS


@pytest_asyncio.fixture
async def admin(db: AsyncSession) -> User:
    email = f"adm-{uuid.uuid4().hex[:6]}@davinci-test.com"
    u = User(open_id=f"email:{email}", email=email, role=UserRole.ADMIN, status=UserStatus.ACTIVE)
    db.add(u)
    await db.commit()
    await db.refresh(u)
    return u


async def _seed_pedido(
    db: AsyncSession,
    admin: User,
    faturador: NfFaturador,
    *,
    numero: str,
    loja: str,
    itens: list[dict],
    total: float | None = None,
    custofrete: float | None = None,
    categoria: str | None = None,
) -> None:
    for i, it in enumerate(itens):
        db.add(
            BlingOrder(
                numero=numero,
                data=datetime(2026, 6, 23, tzinfo=UTC),
                loja=loja,
                situacao="15",
                total=total,
                custofrete=custofrete,
                categoria_nome=it.get("categoria", categoria),
                item_index=i,
                item_codigo=it["sku"],
                item_descricao=it["nome"],
                item_quantidade=it["qtd"],
                itemvalor=it["unit"],
                nome_destinatario="Cleso Menezes",
                cep_destino="30570050",
                endereco_destino="Rua Emídio Beruto",
                numero_destino="30",
                bairro_destino="Cinquentenário",
                cidade_destino="Belo Horizonte",
                uf_destino="MG",
            )
        )


def _col(row: list[str], nome: str) -> str:
    return row[COLUNAS.index(nome)]


@pytest.mark.asyncio
async def test_gera_planilha_avulso_e_exclusivo(
    db: AsyncSession, client: AsyncClient, admin: User, auth_as: Callable[[User | None], None]
):
    auth_as(admin)

    # Faturador avulso (NF cheia, SKU/nome do principal) na loja 1.
    avulso = NfFaturador(
        nome="bling avulso", modo="bling", nf_cheia=True,
        sku_fonte="principal", nome_fonte="produto", ncm="4202.12.10",
    )
    # Faturador exclusivo (0,1%, a001/embalagem) na loja 2.
    exclusivo = NfFaturador(
        nome="bling exclusivo", modo="bling", nf_cheia=False,
        percentual="0.1", sku_fonte="a001", nome_fonte="embalagem",
    )
    db.add_all([avulso, exclusivo])
    await db.flush()

    db.add(StoreInfo(user_id=admin.id, platform="amazon", account_name="l1",
                     bling_store_id="900001", nf_faturador_id=avulso.id))
    db.add(StoreInfo(user_id=admin.id, platform="shopee", account_name="l2",
                     bling_store_id="900002", nf_faturador_id=exclusivo.id))
    await db.flush()

    await _seed_pedido(db, admin, avulso, numero="800001", loja="900001",
                       itens=[{"sku": "dg053.ci", "nome": "Capa Celular", "qtd": 2, "unit": 500}])
    await _seed_pedido(db, admin, exclusivo, numero="800002", loja="900002",
                       itens=[{"sku": "x1", "nome": "Produto X", "qtd": 1, "unit": 1000}])
    await db.commit()

    r = await client.post(
        "/api/nf-cadastro/faturamento/gerar-planilha",
        json={"numeros": ["800001", "800002"]},
    )
    assert r.status_code == 200, r.text
    assert r.headers["X-Pedidos-Ok"] == "2"
    assert r.headers["X-Pedidos-Pulados"] == "0"
    assert "attachment" in r.headers["content-disposition"]

    texto = r.content.decode("utf-8-sig")
    reader = list(csv.reader(io.StringIO(texto), delimiter=";"))
    assert reader[0] == COLUNAS
    linhas = {row[COLUNAS.index("Número pedido")]: row for row in reader[1:]}

    # avulso: SKU/nome do item, valor cheio 500×2 = 1000.
    a = linhas["800001"]
    assert _col(a, "SKU") == "dg053.ci"
    assert _col(a, "Produto") == "Capa Celular"
    assert _col(a, "Valor Total") == "1.000,00"
    assert _col(a, "Cidade Entrega") == "Belo Horizonte"

    # exclusivo: a001/embalagem, 0,1% de 1000 = 1,00.
    e = linhas["800002"]
    assert _col(e, "SKU") == "a001"
    assert _col(e, "Produto") == "embalagem"
    assert _col(e, "Valor Total") == "1,00"


@pytest.mark.asyncio
async def test_faturador_por_tipo_manda_eletro_pro_cheio(
    db: AsyncSession, client: AsyncClient, admin: User, auth_as: Callable[[User | None], None]
):
    """Conta que vende celular E eletro na mesma loja: a regra por tipo
    (store_info.nf_faturador_por_tipo) escolhe o faturador pela categoria do
    item — eletro sai 100% com o SKU real, celular sai 1% como embalagem."""
    auth_as(admin)

    cheio = NfFaturador(
        nome="bling 100%", modo="bling", nf_cheia=True,
        sku_fonte="principal", nome_fonte="produto",
    )
    um_por_cento = NfFaturador(
        nome="bling ex1 1%", modo="bling", nf_cheia=False,
        percentual="1", sku_fonte="a001", nome_fonte="embalagem",
    )
    db.add_all([cheio, um_por_cento])
    await db.flush()

    db.add(StoreInfo(
        user_id=admin.id, platform="ml", account_name="aguiar",
        bling_store_id="950001",
        nf_faturador_id=um_por_cento.id,
        nf_faturador_por_tipo={"eletro": str(cheio.id), "celular": str(um_por_cento.id)},
    ))
    await db.flush()

    await _seed_pedido(db, admin, cheio, numero="850001", loja="950001", categoria="Eletro",
                       itens=[{"sku": "uaf001m1.110", "nome": "airfryer vidro", "qtd": 1, "unit": 384.18}])
    await _seed_pedido(db, admin, um_por_cento, numero="850002", loja="950001", categoria="Celular Kit",
                       itens=[{"sku": "dg053.ci", "nome": "Capa Celular", "qtd": 1, "unit": 1000}])
    # Categoria sem regra própria cai no faturador base da loja (1%).
    await _seed_pedido(db, admin, um_por_cento, numero="850003", loja="950001", categoria="Insumos",
                       itens=[{"sku": "i9", "nome": "Insumo", "qtd": 1, "unit": 200}])
    await db.commit()

    r = await client.post(
        "/api/nf-cadastro/faturamento/gerar-planilha",
        json={"numeros": ["850001", "850002", "850003"]},
    )
    assert r.status_code == 200, r.text
    assert r.headers["X-Pedidos-Ok"] == "3"

    reader = list(csv.reader(io.StringIO(r.content.decode("utf-8-sig")), delimiter=";"))
    linhas = {row[COLUNAS.index("Número pedido")]: row for row in reader[1:]}

    eletro = linhas["850001"]
    assert _col(eletro, "SKU") == "uaf001m1.110"
    assert _col(eletro, "Produto") == "airfryer vidro"
    assert _col(eletro, "Valor Total") == "384,18"

    celular = linhas["850002"]
    assert _col(celular, "SKU") == "a001"
    assert _col(celular, "Valor Total") == "10,00"

    outra = linhas["850003"]
    assert _col(outra, "SKU") == "a001"
    assert _col(outra, "Valor Total") == "2,00"


@pytest.mark.asyncio
async def test_pedido_sem_faturador_vai_pra_pulados(
    db: AsyncSession, client: AsyncClient, admin: User, auth_as: Callable[[User | None], None]
):
    auth_as(admin)
    avulso = NfFaturador(nome="bling avulso", modo="bling", nf_cheia=True,
                         sku_fonte="principal", nome_fonte="produto")
    db.add(avulso)
    await db.flush()
    # loja COM faturador
    db.add(StoreInfo(user_id=admin.id, platform="amazon", account_name="l1",
                     bling_store_id="910001", nf_faturador_id=avulso.id))
    # loja SEM faturador
    db.add(StoreInfo(user_id=admin.id, platform="shopee", account_name="l2",
                     bling_store_id="910002"))
    await db.flush()
    await _seed_pedido(db, admin, avulso, numero="810001", loja="910001",
                       itens=[{"sku": "a", "nome": "A", "qtd": 1, "unit": 100}])
    await _seed_pedido(db, admin, avulso, numero="810002", loja="910002",
                       itens=[{"sku": "b", "nome": "B", "qtd": 1, "unit": 100}])
    await db.commit()

    r = await client.post(
        "/api/nf-cadastro/faturamento/gerar-planilha",
        json={"numeros": ["810001", "810002", "899999"]},
    )
    assert r.status_code == 200, r.text
    assert r.headers["X-Pedidos-Ok"] == "1"
    assert r.headers["X-Pedidos-Pulados"] == "2"
    detalhe = json.loads(base64.b64decode(r.headers["X-Pedidos-Pulados-Detalhe"]))
    motivos = {d["numero"]: d["motivo"] for d in detalhe}
    assert "sem faturador" in motivos["810002"]
    assert "não encontrado" in motivos["899999"]


@pytest.mark.asyncio
async def test_mala_cheia_usa_catalogo_pela_familia(
    db: AsyncSession, client: AsyncClient, admin: User, auth_as: Callable[[User | None], None]
):
    """NF cheia de MALA usa o valor do catálogo pela família do produto (o nome
    'Mala Lisa M2' → abs), NÃO o valor de venda (itemvalor)."""
    auth_as(admin)
    avulso = NfFaturador(
        nome="bling avulso", modo="bling", nf_cheia=True,
        sku_fonte="principal", nome_fonte="produto", ncm="4202.12.10",
    )
    db.add(avulso)
    await db.flush()
    db.add(StoreInfo(user_id=admin.id, platform="amazon", account_name="l1",
                     bling_store_id="920001", nf_faturador_id=avulso.id))
    # Produto com a família no nome (M2 → abs) + catálogo abs/20 = 161.
    db.add(Product(user_id=admin.id, sku="b001.20", name="Mala Lisa M2 tamanho 20 - Roxa"))
    db.add(NfCatalogoMala(modelo="abs", tamanho="20", valor=161))
    await db.flush()
    # Venda a 50 (promocional); a NF deve sair com o cheio 161.
    await _seed_pedido(db, admin, avulso, numero="820001", loja="920001",
                       itens=[{"sku": "b001.20", "nome": "Mala Lisa M2", "qtd": 1, "unit": 50}])
    await db.commit()

    r = await client.post(
        "/api/nf-cadastro/faturamento/gerar-planilha",
        json={"numeros": ["820001"]},
    )
    assert r.status_code == 200, r.text
    texto = r.content.decode("utf-8-sig")
    reader = list(csv.reader(io.StringIO(texto), delimiter=";"))
    a = {row[COLUNAS.index("Número pedido")]: row for row in reader[1:]}["820001"]
    assert _col(a, "Valor Total") == "161,00"


@pytest.mark.asyncio
async def test_percentual_rateia_desconto_do_pedido(
    db: AsyncSession, client: AsyncClient, admin: User, auth_as: Callable[[User | None], None]
):
    """Caso 291422: Shopee põe o desconto no PEDIDO, não no item — o percentual
    tem que incidir sobre o que o cliente PAGOU, não sobre o preço de anúncio.
    Item 600 (anúncio), pago 362,90 → 70% = 254,03 (e não 420,00)."""
    auth_as(admin)
    upseller70 = NfFaturador(
        nome="upseller 70%", modo="upseller", nf_cheia=False,
        percentual="70", sku_fonte="principal", nome_fonte="produto",
    )
    # Mesmo pedido no ML, onde o frete é cobrado À PARTE (pago > soma dos itens):
    # o fator trava em 1 e a NF continua a de hoje.
    exclusivo = NfFaturador(
        nome="bling exclusivo", modo="bling", nf_cheia=False,
        percentual="10", sku_fonte="principal", nome_fonte="produto",
    )
    db.add_all([upseller70, exclusivo])
    await db.flush()
    db.add(StoreInfo(user_id=admin.id, platform="shopee", account_name="poofy",
                     bling_store_id="940001", nf_faturador_id=upseller70.id))
    db.add(StoreInfo(user_id=admin.id, platform="ml", account_name="poofy",
                     bling_store_id="940002", nf_faturador_id=exclusivo.id))
    await db.flush()

    await _seed_pedido(db, admin, upseller70, numero="840001", loja="940001",
                       itens=[{"sku": "b041.18.24", "nome": "Kit Mala", "qtd": 1, "unit": 600}],
                       total=362.90, custofrete=0)
    await _seed_pedido(db, admin, exclusivo, numero="840002", loja="940002",
                       itens=[{"sku": "x1", "nome": "Produto X", "qtd": 1, "unit": 1000}],
                       total=1035.99, custofrete=23.65)
    await db.commit()

    r = await client.post(
        "/api/nf-cadastro/faturamento/gerar-planilha",
        json={"numeros": ["840001", "840002"]},
    )
    assert r.status_code == 200, r.text
    texto = r.content.decode("utf-8-sig")
    reader = list(csv.reader(io.StringIO(texto), delimiter=";"))
    linhas = {row[COLUNAS.index("Número pedido")]: row for row in reader[1:]}
    assert _col(linhas["840001"], "Valor Total") == "254,03"
    assert _col(linhas["840002"], "Valor Total") == "100,00"


@pytest.mark.asyncio
async def test_nenhum_pedido_gerado_422(
    db: AsyncSession, client: AsyncClient, admin: User, auth_as: Callable[[User | None], None]
):
    auth_as(admin)
    r = await client.post(
        "/api/nf-cadastro/faturamento/gerar-planilha",
        json={"numeros": ["000001"]},
    )
    assert r.status_code == 422, r.text
    assert r.json()["detail"]["code"] == "nf_nenhum_pedido_gerado"


def test_chunks_por_limite_bling_e_upseller():
    """Bling limita por PEDIDOS; Upseller pelos DOIS tetos (pedidos E linhas).
    Um pedido nunca é quebrado no meio."""
    from app.services import nf_emissao_gerar as g

    def ped(numero: str, n_linhas: int) -> g._PedidoMontado:
        return g._PedidoMontado(
            numero=numero, faturador_id=uuid.uuid4(), etiqueta_id=None,
            info=None, linhas=[0] * n_linhas
        )

    # Bling: teto 500 pedidos → 501 pedidos de 1 linha viram 2 arquivos.
    grupo = [ped(str(i), 1) for i in range(501)]
    chunks = g._chunks_por_limite(grupo, "bling")
    assert [len(c) for c in chunks] == [500, 1]

    # Upseller: teto 300 pedidos E 1500 linhas. Aqui as LINHAS estouram antes:
    # 300 pedidos de 6 linhas = 1800 > 1500 → corta em 250 (250×6=1500).
    grupo = [ped(str(i), 6) for i in range(300)]
    chunks = g._chunks_por_limite(grupo, "upseller")
    assert [len(c) for c in chunks] == [250, 50]

    # Nunca quebra um pedido: um único pedido com mais linhas que o teto
    # ainda cabe sozinho num arquivo.
    chunks = g._chunks_por_limite([ped("x", 2000)], "upseller")
    assert [len(c) for c in chunks] == [1]

    assert g._chunks_por_limite([], "bling") == []


@pytest.mark.asyncio
async def test_gerar_por_faturador_quebra_em_varios_blocos(
    db: AsyncSession, admin: User, monkeypatch: pytest.MonkeyPatch
):
    """Faturador com mais pedidos que o teto do arquivo gera VÁRIOS blocos
    (comandos), cada um com um subconjunto dos pedidos e nome de arquivo único."""
    from app.services import nf_emissao_gerar as g

    monkeypatch.setattr(g, "_LIMITE_BLING_VENDAS", 2)

    avulso = NfFaturador(
        nome="bling avulso", modo="bling", nf_cheia=True,
        sku_fonte="principal", nome_fonte="produto",
    )
    db.add(avulso)
    await db.flush()
    db.add(StoreInfo(user_id=admin.id, platform="amazon", account_name="l1",
                     bling_store_id="930001", nf_faturador_id=avulso.id))
    await db.flush()
    numeros = ["83001", "83002", "83003"]
    for n in numeros:
        await _seed_pedido(db, admin, avulso, numero=n, loja="930001",
                           itens=[{"sku": "a", "nome": "A", "qtd": 1, "unit": 100}])
    await db.commit()

    res = await g.gerar_por_faturador(db, numeros)
    # 3 pedidos, teto 2 → 2 blocos (2 + 1).
    assert len(res.blocos) == 2
    assert [len(b.numeros) for b in res.blocos] == [2, 1]
    # Cobrem todos os pedidos, sem repetição.
    cobertos = [n for b in res.blocos for n in b.numeros]
    assert sorted(cobertos) == sorted(numeros)
    # Nomes de arquivo únicos.
    nomes = [b.nome_arquivo for b in res.blocos]
    assert len(set(nomes)) == 2
    assert all(nm.endswith(".csv") for nm in nomes)


_DUIMP_MALA = "produto importado pela duimp 26BR0001398015-8"
_DUIMP_AIRFRYER = "produto importado pela duimp 26BR0000876592-9"


def _seed_duimp(db: AsyncSession) -> None:
    """Produtos importados com DUIMP digitada à mão na tela Importação. O SKU é
    gravado COM sufixo (b001.20, uaf001m1.220), igual ao do pedido."""
    db.add_all([
        ImportProduct(categoria="mala", sku="b001.20", custo_bling=Decimal("0"),
                      duimp=_DUIMP_MALA),
        ImportProduct(categoria="eletro", sku="uaf001m1.220", custo_bling=Decimal("0"),
                      duimp=_DUIMP_AIRFRYER),
        # produto SEM duimp — não entra na observação
        ImportProduct(categoria="mala", sku="b026", custo_bling=Decimal("0")),
    ])


@pytest.mark.asyncio
async def test_duimp_vai_na_observacao_do_upseller(db: AsyncSession, admin: User):
    """Shopee (faturador upseller) com mala + airfryer: a DUIMP de cada item
    entra na coluna "Observação" do .xlsx, sem repetir e na ordem dos itens."""
    from app.services import nf_emissao_gerar as g

    fat = NfFaturador(nome="upseller 2%", modo="upseller", nf_cheia=False,
                      percentual="2", sku_fonte="principal", nome_fonte="produto",
                      observacao_duimp=True)
    db.add(fat)
    await db.flush()
    db.add(StoreInfo(user_id=admin.id, platform="shopee", account_name="poofy",
                     bling_store_id="940001", nf_faturador_id=fat.id))
    _seed_duimp(db)
    await db.flush()
    await _seed_pedido(db, admin, fat, numero="84001", loja="940001", itens=[
        {"sku": "b001.20", "nome": "Mala Lisa M2 20", "qtd": 1, "unit": 300},
        {"sku": "uaf001m1.220", "nome": "Air Fryer", "qtd": 1, "unit": 400},
        {"sku": "b026", "nome": "Mala sem duimp", "qtd": 1, "unit": 100},
    ])
    await db.commit()

    res = await g.gerar_por_faturador(db, ["84001"])
    assert len(res.blocos) == 1
    bloco = res.blocos[0]
    assert bloco.nome_arquivo.endswith(".xlsx")

    ws = load_workbook(io.BytesIO(bloco.planilha)).active
    col = nf_upseller._HEADERS.index("Observação") + 1
    esperado = f"{_DUIMP_MALA} | {_DUIMP_AIRFRYER}"
    # dados começam na linha 4; a observação repete em todas as linhas do pedido
    assert [ws.cell(row=r, column=col).value for r in (4, 5, 6)] == [esperado] * 3


@pytest.mark.asyncio
async def test_duimp_vai_na_coluna_observacoes_do_csv(db: AsyncSession, admin: User):
    """Mesma regra no caminho Bling (ML): a DUIMP cai na coluna "Observações"
    do CSV de Importação de Vendas."""
    from app.services import nf_emissao_gerar as g

    fat = NfFaturador(nome="bling avulso", modo="bling", nf_cheia=True,
                      sku_fonte="principal", nome_fonte="produto",
                      observacao_duimp=True)
    db.add(fat)
    await db.flush()
    db.add(StoreInfo(user_id=admin.id, platform="ml", account_name="l1",
                     bling_store_id="940002", nf_faturador_id=fat.id))
    _seed_duimp(db)
    await db.flush()
    await _seed_pedido(db, admin, fat, numero="84002", loja="940002",
                       itens=[{"sku": "b001.20", "nome": "Mala", "qtd": 1, "unit": 300}])
    await db.commit()

    res = await g.gerar_por_faturador(db, ["84002"])
    linhas = list(csv.reader(
        io.StringIO(res.blocos[0].planilha.decode("utf-8-sig")), delimiter=";"
    ))
    assert _col(linhas[1], "Observações") == _DUIMP_MALA


@pytest.mark.asyncio
async def test_sem_flag_duimp_nota_sai_sem_observacao(db: AsyncSession, admin: User):
    """Faturador sem `observacao_duimp` não escreve nada, mesmo com DUIMP
    cadastrada no SKU (a maioria das notas segue assim)."""
    from app.services import nf_emissao_gerar as g

    fat = NfFaturador(nome="bling avulso", modo="bling", nf_cheia=True,
                      sku_fonte="principal", nome_fonte="produto")
    db.add(fat)
    await db.flush()
    db.add(StoreInfo(user_id=admin.id, platform="ml", account_name="l1",
                     bling_store_id="940003", nf_faturador_id=fat.id))
    _seed_duimp(db)
    await db.flush()
    await _seed_pedido(db, admin, fat, numero="84003", loja="940003",
                       itens=[{"sku": "b001.20", "nome": "Mala", "qtd": 1, "unit": 300}])
    await db.commit()

    res = await g.gerar_por_faturador(db, ["84003"])
    linhas = list(csv.reader(
        io.StringIO(res.blocos[0].planilha.decode("utf-8-sig")), delimiter=";"
    ))
    assert _col(linhas[1], "Observações") == ""


def test_extrair_destinatario_nome_vem_do_contato_nao_da_etiqueta():
    """O nome do Comprador tem que casar com o CPF (ambos do `contato`). A
    etiqueta guarda quem RECEBE (Amazon pode mandar outra pessoa) — se virasse
    o nome, a NF sairia com nome que não bate com o CPF."""
    from app.services import nf_emissao_gerar as g
    order = {
        "contato": {"nome": "Leonardo Scorsatto", "numeroDocumento": "03244034101",
                    "tipoPessoa": "F"},
        "transporte": {"etiqueta": {"nome": "Sheila Alves", "endereco": "Quadra CNB 10",
                                    "numero": "301", "municipio": "Brasília",
                                    "uf": "DF", "cep": "72115105"}},
    }
    d = g._extrair_destinatario(order)
    assert d["nome_destinatario"] == "Leonardo Scorsatto"
    assert d["documento"] == "03244034101"
    # Endereço de entrega segue vindo da etiqueta.
    assert d["cidade_destino"] == "Brasília"
    assert d["cep_destino"] == "72115105"


def test_extrair_destinatario_cai_na_etiqueta_sem_contato_nome():
    """Sem nome no contato, usa o da etiqueta como fallback (melhor que vazio)."""
    from app.services import nf_emissao_gerar as g
    order = {
        "contato": {"numeroDocumento": "03244034101"},
        "transporte": {"etiqueta": {"nome": "Sheila Alves"}},
    }
    d = g._extrair_destinatario(order)
    assert d["nome_destinatario"] == "Sheila Alves"
