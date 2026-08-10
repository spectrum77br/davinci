"""Camada de ARQUIVO da emissão pelo UPSELLER (Fase 3a-4).

`nf_upseller.montar_xlsx(loja_nome, pedidos)` monta o .xlsx no template do
Upseller (aba `order_`, 44 colunas): linhas 1-3 são o cabeçalho do modelo e da
linha 4 em diante um item por linha. O destinatário (obrigatório pra NF-e) só
vai na 1ª linha de cada pedido — o Upseller unifica os itens pelo par Nome da
Loja + Nº do Pedido.
"""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import load_workbook

from app.services import nf_upseller
from app.services.nf_emissao import NfLinha
from app.services.nf_relatorio import PedidoInfo

_HEADERS = nf_upseller._HEADERS


def _c(ws, row: int, header: str):
    return ws.cell(row=row, column=_HEADERS.index(header) + 1).value


def _pedido() -> tuple[PedidoInfo, list[NfLinha]]:
    info = PedidoInfo(
        numero="830001",
        data=None,
        nome_destinatario="Cleso Menezes",
        cep_destino="30570050",
        endereco_destino="Rua Emídio Beruto",
        numero_destino="30",
        complemento_destino="Apto 2",
        bairro_destino="Cinquentenário",
        cidade_destino="Belo Horizonte",
        uf_destino="MG",
        documento="12345678901",
        tipo_pessoa="F",
        telefone="31999990000",
    )
    linhas = [
        NfLinha(sku="a001", nome="embalagem", ncm=None,
                quantidade=2, valor_unitario=Decimal("161.00"), valor_total=Decimal("322.00")),
        NfLinha(sku="a002", nome="capa", ncm=None,
                quantidade=1, valor_unitario=Decimal("50.00"), valor_total=Decimal("50.00")),
    ]
    return info, linhas


def test_xlsx_cabecalho_no_template():
    xlsx = nf_upseller.montar_xlsx("Loja Avulsa", [_pedido()])
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    assert ws.title == "order_"
    # linha 3 = os nomes dos campos que o Upseller casa (verbatim, na ordem).
    header = [ws.cell(row=3, column=i + 1).value for i in range(len(_HEADERS))]
    assert header == _HEADERS


def test_xlsx_primeira_linha_traz_destinatario():
    xlsx = nf_upseller.montar_xlsx("Loja Avulsa", [_pedido()])
    ws = load_workbook(io.BytesIO(xlsx)).active
    # dados começam na linha 4 (1ª linha do pedido)
    assert _c(ws, 4, "Nome da Loja*") == "Loja Avulsa"
    assert _c(ws, 4, "Nº do Pedido da Loja*") == "830001"
    assert _c(ws, 4, "Necessita Emitir NF-e*") == "Sim"
    assert _c(ws, 4, "SKU*") == "a001"
    # quantidade/preço são células NUMÉRICAS (texto BR é rejeitado no import)
    assert _c(ws, 4, "Quantidade*") == 2
    assert _c(ws, 4, "Preço Unitário* ") == 161.0
    assert _c(ws, 4, "Método de Pagamento") == "Dinheiro"
    # destinatário (obrigatório pra NF-e) na 1ª linha
    assert _c(ws, 4, "Nome do Destinatário (Obrigatório para NF-e)") == "Cleso Menezes"
    assert _c(ws, 4, "Tipo de Tributação (Obrigatório para NF-e)") == "CPF"
    assert _c(ws, 4, "Número de Tributação (Obrigatório para NF-e)") == "12345678901"
    assert _c(ws, 4, "CEP (Obrigatório para NF-e)") == "30570050"
    # UF vira nome por extenso
    assert _c(ws, 4, "Estado (Obrigatório para NF-e)") == "Minas Gerais"
    assert _c(ws, 4, "Cidade (Obrigatório para NF-e)") == "Belo Horizonte"
    assert _c(ws, 4, "Bairro (Obrigatório para NF-e)") == "Cinquentenário"
    assert _c(ws, 4, "Número") == "30"
    assert _c(ws, 4, "Endereço 1") == "Rua Emídio Beruto"
    assert _c(ws, 4, "Endereço 2") == "Apto 2"
    assert _c(ws, 4, "Nº de Celular") == "31999990000"


def test_xlsx_itens_seguintes_sem_destinatario():
    xlsx = nf_upseller.montar_xlsx("Loja Avulsa", [_pedido()])
    ws = load_workbook(io.BytesIO(xlsx)).active
    # 2ª linha do pedido (linha 5): SKU/qtd/preço presentes, destinatário vazio
    assert _c(ws, 5, "Nome da Loja*") == "Loja Avulsa"
    assert _c(ws, 5, "Nº do Pedido da Loja*") == "830001"
    assert _c(ws, 5, "SKU*") == "a002"
    assert _c(ws, 5, "Quantidade*") == 1
    assert _c(ws, 5, "Preço Unitário* ") == 50.0
    # o bloco do destinatário só na 1ª linha (o Upseller unifica pelo par loja+pedido)
    assert not _c(ws, 5, "Nome do Destinatário (Obrigatório para NF-e)")
    assert not _c(ws, 5, "CEP (Obrigatório para NF-e)")


def test_xlsx_nfe_nao_para_import_ml():
    # Fluxo ML: a NF já saiu do Bling; o Upseller entra só pra puxar a etiqueta,
    # então "Necessita Emitir NF-e = NÃO" (senão o Upseller emitiria 2ª NF). O
    # destinatário continua indo (é quem recebe a etiqueta).
    xlsx = nf_upseller.montar_xlsx("Loja Avulsa", [_pedido()], emitir_nfe=False)
    ws = load_workbook(io.BytesIO(xlsx)).active
    assert _c(ws, 4, "Necessita Emitir NF-e*") == "Não"
    assert _c(ws, 4, "Nome do Destinatário (Obrigatório para NF-e)") == "Cleso Menezes"
    assert _c(ws, 4, "SKU*") == "a001"


def test_usa_catalogo_mala_so_shopee_poofy():
    # O catálogo do Upseller é por CONTA, não por categoria: só a Shopee da
    # poofy importa com m100/m200; as outras lojas da poofy, não.
    assert nf_upseller.usa_catalogo_mala("poofy", "shopee") is True
    assert nf_upseller.usa_catalogo_mala(" Poofy ", "Shopee") is True
    assert nf_upseller.usa_catalogo_mala("poofy", "tiktok") is False
    assert nf_upseller.usa_catalogo_mala("minas", "shopee") is False
    assert nf_upseller.usa_catalogo_mala(None, None) is False


def test_sku_generico_por_categoria():
    # Fora do catálogo de mala é sempre e3, mesmo que a categoria seja mala.
    assert nf_upseller.sku_para_categoria("Celular") == "e3"
    assert nf_upseller.sku_para_categoria("Mala") == "e3"
    assert nf_upseller.sku_para_categoria("Mala Kit") == "e3"
    assert nf_upseller.sku_para_categoria(None) == "e3"
    # No catálogo de mala, kit→m100 e o resto→m200.
    m = {"catalogo_mala": True}
    assert nf_upseller.sku_para_categoria("Mala", **m) == "m200"
    assert nf_upseller.sku_para_categoria("Mala Usada", **m) == "m200"
    assert nf_upseller.sku_para_categoria("Mala Kit", **m) == "m100"
    assert nf_upseller.sku_para_categoria("Celular", **m) == "m200"
    assert nf_upseller.sku_para_categoria(None, **m) == "m200"


def test_skus_para_itens_nao_repete_no_pedido():
    # O Upseller rejeita SKUs duplicados no mesmo pedido: os itens seguintes
    # andam na cadeia da conta (e3→e4→e2→e5) e só então na do outro catálogo.
    assert nf_upseller.skus_para_itens(["Celular", "Celular"]) == ["e3", "e4"]
    assert nf_upseller.skus_para_itens(["Mala", "Mala"]) == ["e3", "e4"]
    # 1 item só: mantém o base.
    assert nf_upseller.skus_para_itens(["Celular"]) == ["e3"]
    assert nf_upseller.skus_para_itens([None]) == ["e3"]
    assert nf_upseller.skus_para_itens(["Celular"] * 4) == ["e3", "e4", "e2", "e5"]
    # 5º item: esgotada a cadeia da conta, segue na do outro catálogo.
    assert nf_upseller.skus_para_itens(["Celular"] * 6) == [
        "e3",
        "e4",
        "e2",
        "e5",
        "m200",
        "m100",
    ]


def test_skus_para_itens_catalogo_mala():
    # Shopee poofy: mala vira m200/m100 e, esgotado, cai na cadeia eletro.
    m = {"catalogo_mala": True}
    assert nf_upseller.skus_para_itens(["Mala", "Mala"], **m) == ["m200", "m100"]
    assert nf_upseller.skus_para_itens(["Mala Kit", "Mala"], **m) == ["m100", "m200"]
    assert nf_upseller.skus_para_itens(["Mala"], **m) == ["m200"]
    assert nf_upseller.skus_para_itens(["Mala"] * 3, **m) == ["m200", "m100", "e3"]
    assert nf_upseller.skus_para_itens(["Mala"] * 7, **m) == [
        "m200",
        "m100",
        "e3",
        "e4",
        "e2",
        "e5",
        "m200",
    ]


def test_loja_upseller_renomeia_victor_mei():
    # A Loja física registrada no Upseller mudou de nome; o arquivo tem que
    # sair com o nome de lá, senão o import rejeita ("loja não existe").
    assert nf_upseller.loja_upseller("victor mei") == "rodgrigues (victor mei)"
    assert nf_upseller.loja_upseller(" Victor Mei ") == "rodgrigues (victor mei)"
    # Conta sem renome passa direto.
    assert nf_upseller.loja_upseller("poofy") == "poofy"
    assert nf_upseller.loja_upseller(None) == ""


def test_loja_avulsa_registrada():
    # A loja do arquivo precisa ser uma loja REGISTRADA na conta Upseller.
    assert nf_upseller.LOJA_AVULSA == "Loja Padrão"


def test_tipo_tributacao_infere_por_documento():
    # sem tipoPessoa, infere pelo tamanho do documento (14 dígitos = CNPJ)
    assert nf_upseller._tipo_tributacao(None, "12345678000199") == "CNPJ"
    assert nf_upseller._tipo_tributacao(None, "12345678901") == "CPF"
    assert nf_upseller._tipo_tributacao("J", "12345678901") == "CNPJ"
    assert nf_upseller._tipo_tributacao(None, None) == ""


def test_estado_por_extenso():
    assert nf_upseller.estado_por_extenso("sp") == "São Paulo"
    assert nf_upseller.estado_por_extenso("MG") == "Minas Gerais"
    # UF desconhecida devolve o valor cru
    assert nf_upseller.estado_por_extenso("ZZ") == "ZZ"
