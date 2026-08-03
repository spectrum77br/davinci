"""Exportar Excel de conferência de um lote de importação.

GET /api/importacao/lotes/{id}/excel → .xlsx com uma linha por produto
com quantidade > 0 nesse lote (modelo/sku/custo/quantidade/total).
"""
from __future__ import annotations

import io
import uuid
from collections.abc import Callable
from datetime import date
from decimal import Decimal

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    ImportLote,
    ImportLoteItem,
    ImportProduct,
    User,
    UserRole,
    UserStatus,
)

PERM = {"importacao": {"view": True, "edit": True, "delete": True}}


@pytest_asyncio.fixture
async def admin_view(db: AsyncSession) -> User:
    u = User(
        open_id=f"email:le-{uuid.uuid4().hex[:6]}@davinci-test.com",
        email=f"le-{uuid.uuid4().hex[:6]}@davinci-test.com",
        role=UserRole.ADMIN,
        status=UserStatus.ACTIVE,
        permissions=PERM,
    )
    db.add(u)
    await db.commit()
    await db.refresh(u)
    return u


def _rows(content: bytes) -> list[tuple]:
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    return [tuple(r) for r in ws.iter_rows(values_only=True)]


@pytest.mark.asyncio
async def test_excel_mala_qtd_vezes_custo(
    db: AsyncSession, client: AsyncClient,
    auth_as: Callable[[User | None], None], admin_view: User,
):
    """Mala: total = quantidade × custo_bling; só produtos com qty > 0."""
    auth_as(admin_view)
    lt = ImportLote(
        id=uuid.uuid4(), categoria="mala", nome="ML25",
        abertura=date(2026, 6, 1),
    )
    db.add(lt)
    p1 = ImportProduct(
        id=uuid.uuid4(), categoria="mala", sku="a075",
        modelo_bling="Chaveiro Poofy", custo_bling=Decimal("9.00"),
    )
    p2 = ImportProduct(
        id=uuid.uuid4(), categoria="mala", sku="a076",
        modelo_bling="Encosto Cabeca", custo_bling=Decimal("18.00"),
    )
    # produto com qty 0 nesse lote → não deve aparecer.
    p3 = ImportProduct(
        id=uuid.uuid4(), categoria="mala", sku="a077",
        modelo_bling="Zerado", custo_bling=Decimal("5.00"),
    )
    db.add_all([p1, p2, p3])
    await db.flush()
    db.add_all([
        ImportLoteItem(id=uuid.uuid4(), lote_id=lt.id, product_id=p1.id, quantidade=5040),
        ImportLoteItem(id=uuid.uuid4(), lote_id=lt.id, product_id=p2.id, quantidade=1680),
        ImportLoteItem(id=uuid.uuid4(), lote_id=lt.id, product_id=p3.id, quantidade=0),
    ])
    await db.commit()

    r = await client.get(f"/api/importacao/lotes/{lt.id}/excel")
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]
    rows = _rows(r.content)
    # header + 2 produtos + total
    assert rows[0] == ("Modelo Bling", "SKU", "Custo Bling", "Quantidade", "Total")
    data = rows[1:-1]
    assert len(data) == 2
    by_sku = {row[1]: row for row in data}
    assert by_sku["a075"] == ("Chaveiro Poofy", "a075", 9.0, 5040, 45360.0)
    assert by_sku["a076"] == ("Encosto Cabeca", "a076", 18.0, 1680, 30240.0)
    # total footer: qtd 6720, valor 75600 (openpyxl lê vazio como None)
    assert rows[-1] == (None, "Total", None, 6720, 75600.0)
    assert "a077" not in by_sku


@pytest.mark.asyncio
async def test_excel_celular_usa_custo_brl(
    db: AsyncSession, client: AsyncClient,
    auth_as: Callable[[User | None], None], admin_view: User,
):
    """Celular: custo unitário = valor_usd × taxa × (1+frete) + adicional."""
    auth_as(admin_view)
    lt = ImportLote(
        id=uuid.uuid4(), categoria="celular", nome="C10",
        abertura=date(2026, 6, 1),
        taxa=Decimal("5"), frete_pct=Decimal("0"), adicional=Decimal("0"),
    )
    db.add(lt)
    p = ImportProduct(
        id=uuid.uuid4(), categoria="celular", sku="i999.sa",
        modelo_bling="Apple Test", custo_bling=Decimal("0"),
    )
    db.add(p)
    await db.flush()
    db.add(ImportLoteItem(
        id=uuid.uuid4(), lote_id=lt.id, product_id=p.id,
        quantidade=10, valor_usd=Decimal("200"),
    ))
    await db.commit()

    r = await client.get(f"/api/importacao/lotes/{lt.id}/excel")
    assert r.status_code == 200, r.text
    rows = _rows(r.content)
    # custoBRL = 200 × 5 = 1000; total = 10 × 1000 = 10000
    assert rows[1] == ("Apple Test", "i999.sa", 1000.0, 10, 10000.0)


@pytest.mark.asyncio
async def test_excel_lote_404(
    db: AsyncSession, client: AsyncClient,
    auth_as: Callable[[User | None], None], admin_view: User,
):
    auth_as(admin_view)
    r = await client.get(f"/api/importacao/lotes/{uuid.uuid4()}/excel")
    assert r.status_code == 404
    assert r.json()["detail"]["code"] == "lote_not_found"


@pytest.mark.asyncio
async def test_excel_exige_auth(
    db: AsyncSession, client: AsyncClient,
    auth_as: Callable[[User | None], None],
):
    auth_as(None)
    r = await client.get(f"/api/importacao/lotes/{uuid.uuid4()}/excel")
    assert r.status_code in (401, 403)
