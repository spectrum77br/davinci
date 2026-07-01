from __future__ import annotations

import pytest
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.models import BlingOrder, Refund, SituacaoBling

pytestmark = pytest.mark.asyncio


def _refund_permissions(*, view: bool = True, edit: bool = True, delete: bool = True) -> dict:
    return {"reembolso": {"view": view, "edit": edit, "delete": delete}}


async def test_list_refunds_requires_view_permission(client, make_user, auth_as):
    user = await make_user(permissions={})
    auth_as(user)

    response = await client.get("/api/refunds")

    assert response.status_code == 403


async def test_create_refund_starts_unchecked_and_can_be_patched(client, make_user, auth_as):
    user = await make_user(permissions=_refund_permissions())
    auth_as(user)

    response = await client.post(
        "/api/refunds",
        json={
            "data": "2026-05-20T12:00:00-03:00",
            "pedido_bling": "123456",
            "pedido_marketplace": "MLB999",
            "plataforma": "ml",
            "conta": "Loja Teste",
            "tipo": "Cliente",
            "prejuizo": 10.5,
            "reembolso": 7.25,
            "chamado_url": "https://www.mercadolivre.com.br/cases/123",
            "chamado_resolvido": False,
        },
    )

    assert response.status_code == 201
    created = response.json()
    assert created["conferido"] is False
    assert created["tipo"] == "Cliente"
    assert created["conta"] == "Loja Teste"
    assert created["chamado_url"] == "https://www.mercadolivre.com.br/cases/123"
    assert created["chamado_resolvido"] is False

    patch = await client.patch(
        f"/api/refunds/{created['id']}",
        json={
            "conferido": True,
            "tipo": "Logistica",
            "chamado": "CH-1",
            "chamado_url": "https://www.mercadolivre.com.br/cases/CH-1",
            "chamado_resolvido": True,
        },
    )

    assert patch.status_code == 200
    updated = patch.json()
    assert updated["conferido"] is True
    assert updated["tipo"] == "Logistica"
    assert updated["chamado"] == "CH-1"
    assert updated["chamado_url"] == "https://www.mercadolivre.com.br/cases/CH-1"
    assert updated["chamado_resolvido"] is True


async def test_cliente_reembolso_is_clamped_to_non_positive(client, make_user, auth_as):
    user = await make_user(permissions=_refund_permissions())
    auth_as(user)

    # Create with Cliente + positive reembolso → auto-negated.
    response = await client.post(
        "/api/refunds",
        json={
            "pedido_bling": "C-1",
            "conta": "Loja X",
            "tipo": "Cliente",
            "reembolso": 12.5,
        },
    )
    assert response.status_code == 201
    created = response.json()
    assert created["tipo"] == "Cliente"
    assert created["reembolso"] == -12.5

    # Patch a Logistica row to tipo=Cliente without resending reembolso →
    # existing positive reembolso must still get clamped.
    logistica = await client.post(
        "/api/refunds",
        json={
            "pedido_bling": "C-2",
            "conta": "Loja X",
            "tipo": "Logistica",
            "reembolso": 8.0,
        },
    )
    assert logistica.status_code == 201
    assert logistica.json()["reembolso"] == 8.0  # not Cliente → untouched

    flip = await client.patch(
        f"/api/refunds/{logistica.json()['id']}",
        json={"tipo": "Cliente"},
    )
    assert flip.status_code == 200
    assert flip.json()["reembolso"] == -8.0

    # Patch only reembolso on an already-Cliente row with a positive value →
    # also clamped.
    bump = await client.patch(
        f"/api/refunds/{created['id']}",
        json={"reembolso": 30.0},
    )
    assert bump.status_code == 200
    assert bump.json()["reembolso"] == -30.0

    # Other tipos accept positive reembolso unchanged.
    other = await client.post(
        "/api/refunds",
        json={
            "pedido_bling": "C-3",
            "conta": "Loja X",
            "tipo": "Extraviado",
            "reembolso": 5.0,
        },
    )
    assert other.status_code == 201
    assert other.json()["reembolso"] == 5.0


async def test_lookup_refund_order_reads_recent_conciliation_view(
    client,
    db: AsyncSession,
    make_user,
    auth_as,
):
    user = await make_user(permissions=_refund_permissions(edit=False, delete=False))
    auth_as(user)
    schema = get_settings().database_schema
    await db.execute(
        text(f'DROP VIEW IF EXISTS "{schema}".vw_conciliacao_margens_marketplace')
    )
    await db.execute(
        text(
            f"""
            CREATE VIEW "{schema}".vw_conciliacao_margens_marketplace AS
            SELECT * FROM (VALUES
                ('2026-05-20T12:00:00+00:00'::timestamptz, '123456'::text, 'MLB999'::text,
                 'ml'::text, NULL::text, 'Conta View'::text, 12.50::numeric),
                ('2026-05-20T12:00:00+00:00'::timestamptz, '123456'::text, 'MLB999'::text,
                 'ml'::text, NULL::text, 'Conta View'::text, 7.25::numeric)
            ) AS t(data, pedido_bling, pedido_marketplace, plataforma_bling,
                   plataforma_financeiro, loja_nome, bling_custo_produtos)
            """  # noqa: S608
        )
    )
    await db.commit()

    response = await client.get("/api/refunds/order-lookup?pedido=MLB999")

    assert response.status_code == 200
    assert response.json() == {
        "items": [
            {
                "data": "2026-05-20T12:00:00Z",
                "pedido_bling": "123456",
                "pedido_marketplace": "MLB999",
                "plataforma": "ml",
                "conta": "Conta View",
                "custo_produto": 19.75,
                "custo_manutencao": None,
            }
        ],
        "historico_disponivel": False,
    }


async def test_lookup_refund_order_surfaces_history_cta_when_recent_view_misses(
    client,
    db: AsyncSession,
    make_user,
    auth_as,
):
    user = await make_user(permissions=_refund_permissions(edit=False, delete=False))
    auth_as(user)
    schema = get_settings().database_schema
    await db.execute(
        text(f'DROP VIEW IF EXISTS "{schema}".vw_conciliacao_margens_marketplace')
    )
    await db.execute(
        text(
            f"""
            CREATE VIEW "{schema}".vw_conciliacao_margens_marketplace AS
            SELECT * FROM (VALUES
                ('2026-05-20T12:00:00+00:00'::timestamptz, '999999'::text, 'MLB999'::text,
                 'ml'::text, NULL::text, 'Conta View'::text, 12.50::numeric)
            ) AS t(data, pedido_bling, pedido_marketplace, plataforma_bling,
                   plataforma_financeiro, loja_nome, bling_custo_produtos)
            WHERE false
            """  # noqa: S608
        )
    )
    await db.execute(
        text(
            """
            INSERT INTO bling_orders (numero, numeroloja)
            VALUES ('123456', 'OLD999')
            """
        )
    )
    await db.commit()

    response = await client.get("/api/refunds/order-lookup?pedido=OLD999")

    assert response.status_code == 200
    assert response.json() == {"items": [], "historico_disponivel": True}


async def test_lookup_refund_order_reads_full_view_when_history_requested(
    client,
    db: AsyncSession,
    make_user,
    auth_as,
):
    user = await make_user(permissions=_refund_permissions(edit=False, delete=False))
    auth_as(user)
    schema = get_settings().database_schema
    await db.execute(
        text(f'DROP VIEW IF EXISTS "{schema}".vw_conciliacao_margens_marketplace')
    )
    await db.execute(
        text(f'DROP VIEW IF EXISTS "{schema}".vw_conciliacao_margens_marketplace_all')
    )
    await db.execute(
        text(
            f"""
            CREATE VIEW "{schema}".vw_conciliacao_margens_marketplace AS
            SELECT * FROM (VALUES
                ('2026-05-20T12:00:00+00:00'::timestamptz, '999999'::text, 'MLB999'::text,
                 'ml'::text, NULL::text, 'Conta View'::text, 12.50::numeric)
            ) AS t(data, pedido_bling, pedido_marketplace, plataforma_bling,
                   plataforma_financeiro, loja_nome, bling_custo_produtos)
            WHERE false
            """  # noqa: S608
        )
    )
    # force_refresh now reads vw_bling_pedidos directly (predicate pushdown),
    # not the heavy vw_conciliacao_margens_marketplace_all view.
    await db.execute(
        text(f'DROP VIEW IF EXISTS "{schema}".vw_bling_pedidos')
    )
    await db.execute(
        text(
            f"""
            CREATE VIEW "{schema}".vw_bling_pedidos AS
            SELECT * FROM (VALUES
                ('2026-04-20T12:00:00+00:00'::timestamptz, '123456'::text, 'OLD999'::text,
                 'shopee'::text, 'Conta Historico'::text, 777::bigint,
                 42.00::numeric, 1::numeric)
            ) AS t(data, numero, numeroloja, marketplace, loja_nome, bling_id,
                   preco_custo, item_quantidade)
            """  # noqa: S608
        )
    )
    await db.execute(
        text(
            """
            INSERT INTO bling_orders (numero, numeroloja)
            VALUES ('123456', 'OLD999')
            """
        )
    )
    await db.commit()

    response = await client.get(
        "/api/refunds/order-lookup?pedido=OLD999&force_refresh=true"
    )

    assert response.status_code == 200
    assert response.json() == {
        "items": [
            {
                "data": "2026-04-20T12:00:00Z",
                "pedido_bling": "123456",
                "pedido_marketplace": "OLD999",
                "plataforma": "shopee",
                "conta": "Conta Historico",
                "custo_produto": 42.0,
                "custo_manutencao": None,
            }
        ],
        "historico_disponivel": False,
    }


async def test_order_cost_sums_bling_custo_produtos(
    client,
    db: AsyncSession,
    make_user,
    auth_as,
):
    user = await make_user(permissions=_refund_permissions(edit=False, delete=False))
    auth_as(user)
    schema = get_settings().database_schema
    await db.execute(text(f'DROP VIEW IF EXISTS "{schema}".vw_conciliacao_margens_marketplace'))
    await db.execute(
        text(
            f"""
            CREATE VIEW "{schema}".vw_conciliacao_margens_marketplace AS
            SELECT * FROM (VALUES
                ('123456'::text, 'Conta View'::text, 12.50::numeric),
                ('123456'::text, 'Conta View'::text, 7.25::numeric),
                ('123456'::text, 'Outra Conta'::text, 99.00::numeric),
                ('999999'::text, 'Conta View'::text, 50.00::numeric)
            ) AS t(pedido_bling, loja_nome, bling_custo_produtos)
            """  # noqa: S608
        )
    )
    await db.commit()

    response = await client.get(
        "/api/refunds/order-cost",
        params={"pedido_bling": "123456", "conta": "Conta View"},
    )

    assert response.status_code == 200
    assert response.json() == {
        "pedido_bling": "123456",
        "conta": "Conta View",
        "custo_produto": 19.75,
    }

    empty = await client.get(
        "/api/refunds/order-cost",
        params={"pedido_bling": "no-match", "conta": "Conta View"},
    )
    assert empty.status_code == 200
    assert empty.json()["custo_produto"] is None


async def test_list_refunds_includes_current_bling_situacao(
    client,
    db: AsyncSession,
    make_user,
    auth_as,
):
    user = await make_user(permissions=_refund_permissions())
    auth_as(user)

    db.add(SituacaoBling(id=9, nome="Atendido"))
    db.add(BlingOrder(numero="555001", situacao="9"))
    # Situacao sem entrada no catalogo -> cai no id cru.
    db.add(BlingOrder(numero="555002", situacao="83965"))
    await db.commit()

    for pedido in ("555001", "555002", "555003"):
        created = await client.post(
            "/api/refunds",
            json={"pedido_bling": pedido, "conta": "Loja Teste"},
        )
        assert created.status_code == 201

    response = await client.get("/api/refunds")
    assert response.status_code == 200
    by_pedido = {item["pedido_bling"]: item for item in response.json()["items"]}

    assert by_pedido["555001"]["situacao_bling"] == "Atendido"
    assert by_pedido["555002"]["situacao_bling"] == "83965"
    # Pedido sem linha em bling_orders -> sem situacao.
    assert by_pedido["555003"]["situacao_bling"] is None


async def test_list_refunds_filters_by_data_range(
    client,
    db: AsyncSession,
    make_user,
    auth_as,
):
    from datetime import datetime
    from zoneinfo import ZoneInfo

    user = await make_user(permissions=_refund_permissions())
    auth_as(user)

    # O filtro vive no GET/_build_where e agora chaveia por conferido_at; insere
    # direto pra não passar pelo refresh de margem do POST. `data` é setada FORA
    # do intervalo de propósito, pra provar que o filtro ignora a coluna Data.
    # "R-B" foi conferido às 23:30 do último dia de maio (SP) e deve entrar com
    # data_fim inclusivo; "R-D" (não conferido, conferido_at NULL) fica de fora.
    sp = ZoneInfo("America/Sao_Paulo")
    fora = datetime(2020, 1, 1, 0, 0, tzinfo=sp)
    db.add_all([
        Refund(pedido_bling="R-A", conta="Loja Teste", data=fora, conferido=True,
               conferido_at=datetime(2026, 5, 10, 9, 0, tzinfo=sp)),
        Refund(pedido_bling="R-B", conta="Loja Teste", data=fora, conferido=True,
               conferido_at=datetime(2026, 5, 31, 23, 30, tzinfo=sp)),
        Refund(pedido_bling="R-C", conta="Loja Teste", data=fora, conferido=True,
               conferido_at=datetime(2026, 6, 2, 9, 0, tzinfo=sp)),
        Refund(pedido_bling="R-D", conta="Loja Teste", data=fora, conferido=False,
               conferido_at=None),
    ])
    await db.commit()

    def _pedidos(resp):
        return {item["pedido_bling"] for item in resp.json()["items"]}

    # Intervalo fechado de maio: pega A e B, exclui C (junho) e D (não conferido).
    resp = await client.get("/api/refunds?data_inicio=2026-05-01&data_fim=2026-05-31")
    assert resp.status_code == 200
    assert _pedidos(resp) == {"R-A", "R-B"}

    # Só limite inferior: A, B e C; exclui apenas o não conferido.
    resp = await client.get("/api/refunds?data_inicio=2026-05-01")
    assert resp.status_code == 200
    assert _pedidos(resp) == {"R-A", "R-B", "R-C"}

    # Só limite superior: exclui C (junho) e o não conferido.
    resp = await client.get("/api/refunds?data_fim=2026-05-31")
    assert resp.status_code == 200
    assert _pedidos(resp) == {"R-A", "R-B"}


async def test_export_refunds_xlsx_respects_filters(
    client,
    db: AsyncSession,
    make_user,
    auth_as,
):
    from io import BytesIO

    from openpyxl import load_workbook

    user = await make_user(permissions=_refund_permissions())
    auth_as(user)

    db.add(SituacaoBling(id=9, nome="Atendido"))
    db.add(BlingOrder(numero="777001", situacao="9"))
    await db.commit()

    for pedido, tipo in (("777001", "Cliente"), ("777002", "Logistica")):
        created = await client.post(
            "/api/refunds",
            json={"pedido_bling": pedido, "conta": "Loja Teste", "tipo": tipo},
        )
        assert created.status_code == 201

    response = await client.get("/api/refunds/export.xlsx?tipo=Cliente")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(BytesIO(response.content))
    rows = list(wb.active.iter_rows(values_only=True))
    header = rows[0]
    assert header[0] == "Data"
    assert "Situação Bling" in header

    data_rows = rows[1:]
    assert len(data_rows) == 1
    row = dict(zip(header, data_rows[0], strict=True))
    assert row["Pedido Bling"] == "777001"
    assert row["Tipo"] == "Cliente"
    assert row["Situação Bling"] == "Atendido"
    assert row["Conferido"] == "Não"


async def test_export_refunds_xlsx_requires_view_permission(client, make_user, auth_as):
    user = await make_user(permissions={})
    auth_as(user)

    response = await client.get("/api/refunds/export.xlsx")

    assert response.status_code == 403
