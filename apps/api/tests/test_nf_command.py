"""Outbox da importação avulsa (Fase 3a-4) + superfície M2M do executor.

`POST /api/nf-cadastro/faturamento/enfileirar` gera a planilha POR FATURADOR e
cria um NfCommand (CSV congelado) por faturador, marcando cada pedido como
'processando' em nf_faturamento. O executor local faz poll de
`/agent/lease` (X-Agent-Token), pega o login do faturador + a planilha em
base64, e reporta em `/agent/commands/{id}/result` (done → 'ok', failed →
'erro').
"""

from __future__ import annotations

import base64
import io
import uuid
from collections.abc import Callable
from datetime import UTC, datetime

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.models import (
    BlingOrder,
    NfCommand,
    NfEtiqueta,
    NfFaturador,
    NfFaturamento,
    Product,
    StoreInfo,
    User,
    UserRole,
    UserStatus,
)
from app.security.cipher import encrypt
from app.services import nf_emissao_gerar, nf_relatorio

_TOKEN = "nf-agent-test-token"


async def _async_return(value: object) -> object:
    """`_bling_client_opt` é async — o monkeypatch devolve a coroutine pronta."""
    return value


@pytest_asyncio.fixture
async def admin(db: AsyncSession) -> User:
    email = f"adm-{uuid.uuid4().hex[:6]}@davinci-test.com"
    u = User(open_id=f"email:{email}", email=email, role=UserRole.ADMIN, status=UserStatus.ACTIVE)
    db.add(u)
    await db.commit()
    await db.refresh(u)
    return u


async def _seed_pedido(
    db: AsyncSession, *, numero: str, loja: str, sku: str, nome: str, unit: float,
    bling_id: int = 700001, numeroloja: str | None = None,
) -> None:
    db.add(
        BlingOrder(
            bling_id=bling_id,
            numero=numero,
            numeroloja=numeroloja,
            data=datetime(2026, 6, 23, tzinfo=UTC),
            loja=loja,
            situacao="15",
            item_index=0,
            item_codigo=sku,
            item_descricao=nome,
            item_quantidade=1,
            itemvalor=unit,
            nome_destinatario="Cleso Menezes",
            cep_destino="30570050",
            endereco_destino="Rua Emídio Beruto",
            numero_destino="30",
            bairro_destino="Cinquentenário",
            cidade_destino="Belo Horizonte",
            uf_destino="MG",
        )
    )


async def _seed_dois_faturadores(db: AsyncSession, admin: User) -> None:
    """Duas lojas com faturadores distintos + um pedido cada → 2 comandos."""
    f1 = NfFaturador(
        nome="bling avulso", modo="bling", nf_cheia=True,
        sku_fonte="principal", nome_fonte="produto", ncm="4202.12.10",
        ads_power="perfil-A", usuario="user-a", senha_enc=encrypt("segredoA"),
    )
    f2 = NfFaturador(
        nome="bling exclusivo", modo="bling", nf_cheia=False,
        percentual="0.1", sku_fonte="a001", nome_fonte="embalagem",
        ads_power="perfil-B", usuario="user-b", senha_enc=encrypt("segredoB"),
    )
    db.add_all([f1, f2])
    await db.flush()
    db.add(StoreInfo(user_id=admin.id, platform="amazon", account_name="l1",
                     bling_store_id="930001", nf_faturador_id=f1.id))
    db.add(StoreInfo(user_id=admin.id, platform="shopee", account_name="l2",
                     bling_store_id="930002", nf_faturador_id=f2.id))
    await db.flush()
    await _seed_pedido(db, numero="830001", loja="930001", sku="dg053.ci", nome="Capa", unit=500)
    await _seed_pedido(db, numero="830002", loja="930002", sku="x1", nome="Produto X", unit=1000)
    await db.commit()


@pytest.mark.asyncio
async def test_enfileirar_cria_comando_por_faturador(
    db: AsyncSession, client: AsyncClient, admin: User, auth_as: Callable[[User | None], None]
):
    auth_as(admin)
    await _seed_dois_faturadores(db, admin)

    r = await client.post(
        "/api/nf-cadastro/faturamento/enfileirar",
        json={"numeros": ["830001", "830002"]},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["comandos"] == 2
    assert body["pedidos_ok"] == 2
    assert body["pulados"] == []

    cmds = (await db.execute(select(NfCommand))).scalars().all()
    assert len(cmds) == 2
    assert {c.status for c in cmds} == {"pending"}
    # modo 'bling' congela o CSV do relatório de vendas (tela Importar vendas
    # aceita CSV); o arquivo começa com o BOM UTF-8
    assert all(c.planilha and c.nome_arquivo.endswith(".csv") for c in cmds)
    assert all(c.planilha.startswith(b"\xef\xbb\xbf") for c in cmds)
    # cada comando cobre exatamente o pedido da sua loja/faturador
    cobertos = sorted(n for c in cmds for n in c.numeros)
    assert cobertos == ["830001", "830002"]

    # cada pedido virou 'processando' em nf_faturamento
    fats = {f.pedido_bling: f for f in (await db.execute(select(NfFaturamento))).scalars().all()}
    assert fats["830001"].status_faturamento == "processando"
    assert fats["830002"].status_faturamento == "processando"


@pytest.mark.asyncio
async def test_enfileirar_sem_faturador_422(
    db: AsyncSession, client: AsyncClient, admin: User, auth_as: Callable[[User | None], None]
):
    auth_as(admin)
    # loja SEM faturador → nenhum bloco gerado
    db.add(StoreInfo(user_id=admin.id, platform="shopee", account_name="l2",
                     bling_store_id="940002"))
    await db.flush()
    await _seed_pedido(db, numero="840002", loja="940002", sku="b", nome="B", unit=100)
    await db.commit()

    r = await client.post(
        "/api/nf-cadastro/faturamento/enfileirar",
        json={"numeros": ["840002"]},
    )
    assert r.status_code == 422, r.text
    assert r.json()["detail"]["code"] == "nf_nenhum_pedido_gerado"


class _FakeBlingSituacao:
    """Captura os PATCH de situação que o motor manda pro Bling."""

    def __init__(self) -> None:
        self.chamadas: list[tuple[int, int]] = []

    async def update_order_situacao(self, bling_order_id: int, situacao_id: int) -> None:
        self.chamadas.append((bling_order_id, situacao_id))


@pytest.mark.asyncio
async def test_enfileirar_pula_pedido_sem_estoque(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """Item com saldo virtual negativo não vira NF/etiqueta: o pedido sai da fila
    e vai pra Aguardando Cancelamento (83955) no Bling."""
    auth_as(admin)
    await _seed_dois_faturadores(db, admin)
    db.add(Product(user_id=admin.id, sku="x1", name="Produto X", stock=-3))
    db.add(Product(user_id=admin.id, sku="dg053.ci", name="Capa", stock=5))
    await db.commit()
    fake = _FakeBlingSituacao()
    monkeypatch.setattr(
        nf_emissao_gerar, "_bling_client_opt", lambda _s: _async_return(fake)
    )

    r = await client.post(
        "/api/nf-cadastro/faturamento/enfileirar",
        json={"numeros": ["830001", "830002"]},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["comandos"] == 1
    assert body["pedidos_ok"] == 1
    assert [p["numero"] for p in body["pulados"]] == ["830002"]
    assert "sem estoque" in body["pulados"][0]["motivo"]

    # só o pedido COM estoque foi enfileirado
    cmds = (await db.execute(select(NfCommand))).scalars().all()
    assert [n for c in cmds for n in c.numeros] == ["830001"]

    # o sem estoque foi pra Aguardando Cancelamento no Bling e localmente
    assert fake.chamadas == [(700001, 83955)]
    situacoes = {
        r.numero: r.situacao
        for r in (
            await db.execute(select(BlingOrder.numero, BlingOrder.situacao))
        ).all()
    }
    assert situacoes["830002"] == "83955"
    assert situacoes["830001"] == "15"

    fats = {f.pedido_bling: f for f in (await db.execute(select(NfFaturamento))).scalars().all()}
    assert fats["830002"].status_faturamento == "sem_estoque"
    assert fats["830001"].status_faturamento == "processando"


@pytest.mark.asyncio
async def test_enfileirar_todos_sem_estoque_422(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """Nenhum pedido sobra pra fila → 422, mas os pulados por estoque aparecem
    na resposta (o front lista o motivo)."""
    auth_as(admin)
    await _seed_dois_faturadores(db, admin)
    db.add(Product(user_id=admin.id, sku="x1", name="Produto X", stock=-1))
    await db.commit()
    monkeypatch.setattr(
        nf_emissao_gerar,
        "_bling_client_opt",
        lambda _s: _async_return(_FakeBlingSituacao()),
    )

    r = await client.post(
        "/api/nf-cadastro/faturamento/enfileirar", json={"numeros": ["830002"]}
    )
    assert r.status_code == 422, r.text
    detail = r.json()["detail"]
    assert detail["code"] == "nf_nenhum_pedido_gerado"
    assert [p["numero"] for p in detail["pulados"]] == ["830002"]


@pytest.mark.asyncio
async def test_enfileirar_saldo_zero_tem_estoque(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None],
):
    """Saldo virtual 0 já desconta a reserva do próprio pedido — ainda é
    atendível. Só o NEGATIVO bloqueia."""
    auth_as(admin)
    await _seed_dois_faturadores(db, admin)
    db.add(Product(user_id=admin.id, sku="x1", name="Produto X", stock=0))
    await db.commit()

    r = await client.post(
        "/api/nf-cadastro/faturamento/enfileirar", json={"numeros": ["830002"]}
    )
    assert r.status_code == 200, r.text
    assert r.json()["pedidos_ok"] == 1


@pytest.mark.asyncio
async def test_agent_lease_entrega_login_e_planilha(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_dois_faturadores(db, admin)
    await client.post("/api/nf-cadastro/faturamento/enfileirar",
                      json={"numeros": ["830001", "830002"]})

    r = await client.post(
        "/api/nf-cadastro/agent/lease",
        json={"limit": 5},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 200, r.text
    cmds = r.json()["commands"]
    assert len(cmds) == 2
    por_user = {c["usuario"]: c for c in cmds}
    a = por_user["user-a"]
    assert a["ads_power"] == "perfil-A"
    assert a["senha"] == "segredoA"  # descriptografada no lease
    assert a["numeros"] == ["830001"]
    assert a["ncm"] == "4202.12.10"  # NCM do faturador entregue pro agente
    assert por_user["user-b"]["ncm"] is None  # faturador sem NCM
    # planilha em base64 decodifica pro CSV congelado (BOM UTF-8 + cabeçalho)
    csv_bytes = base64.b64decode(a["planilha_b64"])
    assert csv_bytes.startswith(b"\xef\xbb\xbf")
    assert b"N\xc3\xbamero pedido" in csv_bytes

    # os comandos viraram 'claimed' com attempts=1
    rows = (await db.execute(select(NfCommand))).scalars().all()
    assert {c.status for c in rows} == {"claimed"}
    assert all(c.attempts == 1 and c.claimed_at is not None for c in rows)


@pytest.mark.asyncio
async def test_agent_lease_filtra_por_action(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    """`actions` reivindica só as actions pedidas; `exclude_actions` pula as
    excluídas (loop contínuo vs passe horário de etiquetas)."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    db.add_all([
        NfCommand(action="import_avulsa", numeros=["910001"], planilha=b"x",
                  nome_arquivo="a.csv", status="pending", source="manual"),
        NfCommand(action="imprimir_etiqueta", numeros=["910002"], planilha=b"",
                  nome_arquivo="", status="pending", source="auto"),
    ])
    await db.commit()

    # passe horário: só etiquetas
    r = await client.post(
        "/api/nf-cadastro/agent/lease",
        json={"limit": 5, "actions": ["imprimir_etiqueta"]},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 200, r.text
    cmds = r.json()["commands"]
    assert [c["action"] for c in cmds] == ["imprimir_etiqueta"]

    # loop contínuo: tudo MENOS etiquetas (o import segue pending)
    r = await client.post(
        "/api/nf-cadastro/agent/lease",
        json={"limit": 5, "exclude_actions": ["imprimir_etiqueta"]},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 200, r.text
    cmds = r.json()["commands"]
    assert [c["action"] for c in cmds] == ["import_avulsa"]

    # sem filtro segue funcionando (compat) — nada pending sobrou
    r = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                          headers={"X-Agent-Token": _TOKEN})
    assert r.json()["commands"] == []


@pytest.mark.asyncio
async def test_agent_lease_sem_token_401(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    # sem header
    r = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5})
    assert r.status_code == 401
    # token errado
    r = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                          headers={"X-Agent-Token": "errado"})
    assert r.status_code == 401


@pytest.mark.asyncio
async def test_agent_lease_token_vazio_fecha(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    # token não configurado no servidor → endpoint fechado mesmo com header
    monkeypatch.setattr(get_settings(), "nf_agent_token", "")
    r = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                          headers={"X-Agent-Token": "qualquer"})
    assert r.status_code == 401


@pytest.mark.asyncio
async def test_agent_result_done_marca_ok(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_dois_faturadores(db, admin)
    await client.post("/api/nf-cadastro/faturamento/enfileirar",
                      json={"numeros": ["830001", "830002"]})
    lease = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                              headers={"X-Agent-Token": _TOKEN})
    cmd = lease.json()["commands"][0]

    r = await client.post(
        f"/api/nf-cadastro/agent/commands/{cmd['id']}/result",
        json={"status": "done"},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "done"

    row = await db.get(NfCommand, uuid.UUID(cmd["id"]))
    await db.refresh(row)
    assert row.status == "done"
    assert row.completed_at is not None

    numero = cmd["numeros"][0]
    fat = (await db.execute(
        select(NfFaturamento).where(NfFaturamento.pedido_bling == numero)
    )).scalar_one()
    await db.refresh(fat)
    assert fat.status_faturamento == "ok"
    assert fat.erro_faturamento is None


@pytest.mark.asyncio
async def test_agent_result_failed_marca_erro(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_dois_faturadores(db, admin)
    await client.post("/api/nf-cadastro/faturamento/enfileirar",
                      json={"numeros": ["830001", "830002"]})
    lease = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                              headers={"X-Agent-Token": _TOKEN})
    cmd = lease.json()["commands"][0]

    r = await client.post(
        f"/api/nf-cadastro/agent/commands/{cmd['id']}/result",
        json={"status": "failed", "result": "AdsPower não abriu"},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "failed"

    numero = cmd["numeros"][0]
    fat = (await db.execute(
        select(NfFaturamento).where(NfFaturamento.pedido_bling == numero)
    )).scalar_one()
    await db.refresh(fat)
    assert fat.status_faturamento == "erro"
    assert fat.erro_faturamento == "AdsPower não abriu"


@pytest.mark.asyncio
async def test_agent_result_comando_inexistente_404(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    r = await client.post(
        f"/api/nf-cadastro/agent/commands/{uuid.uuid4()}/result",
        json={"status": "done"},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 404
    assert r.json()["detail"]["code"] == "nf_command_not_found"


@pytest.mark.asyncio
async def test_enfileirar_upseller_gera_xlsx(
    db: AsyncSession, client: AsyncClient, admin: User, auth_as: Callable[[User | None], None]
):
    """Faturador modo 'upseller' congela a planilha no template .xlsx do Upseller
    (não o CSV do Bling), servida com o media type de planilha."""
    auth_as(admin)
    f = NfFaturador(
        nome="upseller 100%", modo="upseller", nf_cheia=True,
        sku_fonte="principal", nome_fonte="produto", ncm="4202.12.10",
        ads_power="perfil-U", usuario="user-u", senha_enc=encrypt("segredoU"),
    )
    db.add(f)
    await db.flush()
    db.add(StoreInfo(user_id=admin.id, platform="amazon", account_name="lU",
                     bling_store_id="950001", nf_faturador_id=f.id))
    await db.flush()
    await _seed_pedido(db, numero="850001", loja="950001", sku="dg053.ci", nome="Capa", unit=500)
    await db.commit()

    r = await client.post(
        "/api/nf-cadastro/faturamento/enfileirar",
        json={"numeros": ["850001"]},
    )
    assert r.status_code == 200, r.text
    assert r.json()["comandos"] == 1

    cmd = (await db.execute(select(NfCommand))).scalars().one()
    assert cmd.nome_arquivo.endswith(".xlsx")
    # .xlsx é um zip (assinatura PK\x03\x04), não o BOM do CSV
    assert cmd.planilha.startswith(b"PK\x03\x04")

    # "Nome da Loja*" = a CONTA de marketplace do pedido (cada conta é uma Loja
    # registrada no Upseller), não a genérica "Loja Padrão".
    from app.services.nf_upseller import _HEADERS
    ws = load_workbook(io.BytesIO(cmd.planilha)).active
    col_loja = _HEADERS.index("Nome da Loja*") + 1
    assert ws.cell(row=4, column=col_loja).value == "lU"


@pytest.mark.asyncio
async def test_agent_planilha_serve_zip(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_dois_faturadores(db, admin)
    await client.post("/api/nf-cadastro/faturamento/enfileirar",
                      json={"numeros": ["830001", "830002"]})
    cmd = (await db.execute(select(NfCommand))).scalars().first()

    r = await client.get(
        f"/api/nf-cadastro/agent/commands/{cmd.id}/planilha",
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 200, r.text
    assert "attachment" in r.headers["content-disposition"]
    # modo 'bling' serve o CSV (BOM UTF-8) na tela Importar vendas
    assert r.content.startswith(b"\xef\xbb\xbf")
    assert r.headers["content-type"].startswith(nf_relatorio.CSV_MEDIA)


def _etiqueta_pdf() -> bytes:
    """Etiqueta sintética com REMETENTE/DESTINATÁRIO + bloco NF pra transformar."""
    import fitz

    doc = fitz.open()
    page = doc.new_page(width=300, height=442)
    page.insert_text((20, 20), "DESTINATÁRIO", fontsize=7, fontname="helv")
    page.insert_text((20, 32), "Fulano De Tal", fontsize=8, fontname="helv")
    page.insert_text((20, 280), "REMETENTE", fontsize=7, fontname="helv")
    page.insert_text((20, 294), "Loja Origem XYZ", fontsize=8, fontname="helv")
    page.insert_text((20, 340), "DANFE SIMPLIFICADO", fontsize=6, fontname="helv")
    page.insert_text(
        (20, 350), "35260730734713000140550040000029931203736630",
        fontsize=5, fontname="helv",
    )
    return doc.tobytes()


@pytest.mark.asyncio
async def test_agent_etiqueta_upsert_transformada(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    """A marionete sobe o PDF cru; o endpoint transforma (remetente=destinatário,
    sem bloco NF) e faz upsert em nf_etiqueta_arquivo; a 2ª subida regrava."""
    import fitz

    from app.models import NfEtiquetaArquivo

    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    r = await client.post(
        "/api/nf-cadastro/agent/etiqueta",
        data={"pedido_bling": "870001"},
        files={"file": ("etq.pdf", _etiqueta_pdf(), "application/pdf")},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 200, r.text
    assert r.json()["pedido_bling"] == "870001"

    row = (
        await db.execute(
            select(NfEtiquetaArquivo).where(
                NfEtiquetaArquivo.pedido_bling == "870001"
            )
        )
    ).scalar_one()
    txt = fitz.open(stream=row.blob, filetype="pdf")[0].get_text()
    assert "Loja Origem" not in txt  # remetente virou destinatário
    assert "DANFE" not in txt  # bloco NF removido
    assert txt.count("Fulano") == 2

    # 2ª subida (upsert) não duplica e regrava
    r2 = await client.post(
        "/api/nf-cadastro/agent/etiqueta",
        data={"pedido_bling": "870001", "destinatario_nome": "Beltrano Oficial"},
        files={"file": ("etq.pdf", _etiqueta_pdf(), "application/pdf")},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r2.status_code == 200, r2.text
    db.expire_all()  # o request comitou em outra sessão; recarrega do banco
    rows = (
        await db.execute(
            select(NfEtiquetaArquivo).where(
                NfEtiquetaArquivo.pedido_bling == "870001"
            )
        )
    ).scalars().all()
    assert len(rows) == 1
    txt2 = fitz.open(stream=rows[0].blob, filetype="pdf")[0].get_text()
    assert "Beltrano Oficial" in txt2


@pytest.mark.asyncio
async def test_agent_nf_grava_danfe_em_linha_existente(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    """A etiqueta já chegou; a marionete sobe o DANFE do Bling → grava nf_pdf
    sem tocar o blob da etiqueta transformada."""
    from app.models import NfEtiquetaArquivo

    db.add(NfEtiquetaArquivo(
        pedido_bling="880001",
        filename="etiqueta_880001.pdf",
        content_type="application/pdf",
        size_bytes=3,
        blob=b"ETQ",
    ))
    await db.commit()

    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    r = await client.post(
        "/api/nf-cadastro/agent/nf",
        data={"pedido_bling": "880001"},
        files={"file": ("danfe.pdf", b"%PDF-DANFE-BYTES", "application/pdf")},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 200, r.text
    assert r.json()["nf_size_bytes"] == len(b"%PDF-DANFE-BYTES")

    db.expire_all()
    row = (
        await db.execute(
            select(NfEtiquetaArquivo).where(
                NfEtiquetaArquivo.pedido_bling == "880001"
            )
        )
    ).scalar_one()
    assert row.blob == b"ETQ"  # etiqueta intacta
    assert row.nf_pdf == b"%PDF-DANFE-BYTES"
    assert row.nf_size_bytes == len(b"%PDF-DANFE-BYTES")


@pytest.mark.asyncio
async def test_agent_nf_cria_linha_quando_etiqueta_nao_chegou(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    """A NF pode chegar antes da etiqueta: cria a linha só com a NF (blob vazio)."""
    from app.models import NfEtiquetaArquivo

    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    r = await client.post(
        "/api/nf-cadastro/agent/nf",
        data={"pedido_bling": "880002"},
        files={"file": ("danfe.pdf", b"%PDF-DANFE", "application/pdf")},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 200, r.text

    db.expire_all()
    row = (
        await db.execute(
            select(NfEtiquetaArquivo).where(
                NfEtiquetaArquivo.pedido_bling == "880002"
            )
        )
    ).scalar_one()
    assert row.blob == b""  # sem etiqueta ainda
    assert row.nf_pdf == b"%PDF-DANFE"


@pytest.mark.asyncio
async def test_agent_nf_vazia_400(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    r = await client.post(
        "/api/nf-cadastro/agent/nf",
        data={"pedido_bling": "880003"},
        files={"file": ("danfe.pdf", b"", "application/pdf")},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 400, r.text
    assert r.json()["detail"]["code"] == "nf_pdf_vazia"


@pytest.mark.asyncio
async def test_agent_nf_sem_token_401(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    r = await client.post(
        "/api/nf-cadastro/agent/nf",
        data={"pedido_bling": "880004"},
        files={"file": ("danfe.pdf", b"%PDF", "application/pdf")},
    )
    assert r.status_code == 401, r.text


async def _seed_upseller(db: AsyncSession, admin: User, *, numero: str = "850001") -> None:
    """Faturador Upseller + 1 pedido → 1 comando .xlsx (fluxo com etiqueta)."""
    f = NfFaturador(
        nome="upseller 100%", modo="upseller", nf_cheia=True,
        sku_fonte="principal", nome_fonte="produto", ncm="4202.12.10",
        ads_power="perfil-U", usuario="user-u", senha_enc=encrypt("segredoU"),
    )
    db.add(f)
    await db.flush()
    db.add(StoreInfo(user_id=admin.id, platform="amazon", account_name="lU",
                     bling_store_id="950001", nf_faturador_id=f.id))
    await db.flush()
    await _seed_pedido(db, numero=numero, loja="950001", sku="dg053.ci", nome="Capa",
                       unit=500, numeroloja=f"PLAT-{numero}")
    await db.commit()


async def _import_done(client: AsyncClient, numeros: list[str]) -> None:
    """Enfileira o faturamento, reivindica e reporta 'done' (dispara o motor)."""
    await client.post("/api/nf-cadastro/faturamento/enfileirar", json={"numeros": numeros})
    lease = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                              headers={"X-Agent-Token": _TOKEN})
    imp = next(c for c in lease.json()["commands"] if c["action"] == "import_avulsa")
    await client.post(f"/api/nf-cadastro/agent/commands/{imp['id']}/result",
                      json={"status": "done"}, headers={"X-Agent-Token": _TOKEN})


async def _emissao_done(client: AsyncClient, *, status: str = "done") -> None:
    """Reivindica os comandos da cadeia NF do Upseller e reporta o resultado.

    Importar o avulso só põe o pedido na fila "Para Emitir": a etiqueta só fica
    imprimível depois que a NF é emitida, o XML exportado e subido no pedido
    original — por isso a etiqueta nasce daqui, não do import."""
    lease = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 20},
                              headers={"X-Agent-Token": _TOKEN})
    for cmd in lease.json()["commands"]:
        if cmd["action"] != "emitir_nf_upseller":
            continue
        await client.post(f"/api/nf-cadastro/agent/commands/{cmd['id']}/result",
                          json={"status": status}, headers={"X-Agent-Token": _TOKEN})


async def _ciclo_upseller(client: AsyncClient, numeros: list[str]) -> None:
    """Fluxo Upseller completo até a etiqueta ser enfileirada: importa o avulso e
    fecha a cadeia NF (emitir → exportar XML → subir no original)."""
    await _import_done(client, numeros)
    await _emissao_done(client)


@pytest.mark.asyncio
async def test_import_upseller_auto_enfileira_emissao(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """Faturador Upseller: importar o avulso NÃO emite a nota — o motor encadeia
    UM comando `emitir_nf_upseller` por pedido e o faturamento segue
    'processando' (a etiqueta só nasce quando a emissão fechar)."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_upseller(db, admin)
    await _import_done(client, ["850001"])

    emi = (await db.execute(
        select(NfCommand).where(NfCommand.action == "emitir_nf_upseller")
    )).scalars().all()
    assert len(emi) == 1
    assert emi[0].numeros == ["850001"]
    assert emi[0].status == "pending"
    assert emi[0].source == "auto"
    assert emi[0].planilha == b""
    assert emi[0].ads_power == "perfil-U"

    etq = (await db.execute(
        select(NfCommand).where(NfCommand.action == "imprimir_etiqueta")
    )).scalars().all()
    assert etq == []

    fat = (await db.execute(
        select(NfFaturamento).where(NfFaturamento.pedido_bling == "850001")
    )).scalar_one()
    await db.refresh(fat)
    assert fat.status_faturamento == "processando"


@pytest.mark.asyncio
async def test_faturamento_ok_auto_enfileira_etiqueta(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """Cadeia NF fechada 'ok' → o faturamento fecha e o motor auto-enfileira UM
    comando `imprimir_etiqueta` por pedido (planilha vazia) + marca
    status_etiqueta='processando'."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_upseller(db, admin)
    await _ciclo_upseller(client, ["850001"])

    etq = (await db.execute(
        select(NfCommand).where(NfCommand.action == "imprimir_etiqueta")
    )).scalars().all()
    assert len(etq) == 1
    assert etq[0].numeros == ["850001"]
    assert etq[0].status == "pending"
    assert etq[0].source == "auto"
    assert etq[0].planilha == b""
    assert etq[0].faturador_id is not None

    fat = (await db.execute(
        select(NfFaturamento).where(NfFaturamento.pedido_bling == "850001")
    )).scalar_one()
    await db.refresh(fat)
    assert fat.status_faturamento == "ok"
    assert fat.status_etiqueta == "processando"


@pytest.mark.asyncio
async def test_emissao_falhou_marca_faturamento_erro(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """Cadeia NF que falha na tela deixa o faturamento em 'erro' e NÃO encadeia a
    etiqueta (sem NF no pedido original não há o que imprimir)."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_upseller(db, admin)
    await _import_done(client, ["850001"])
    await _emissao_done(client, status="failed")

    etq = (await db.execute(
        select(NfCommand).where(NfCommand.action == "imprimir_etiqueta")
    )).scalars().all()
    assert etq == []

    fat = (await db.execute(
        select(NfFaturamento).where(NfFaturamento.pedido_bling == "850001")
    )).scalar_one()
    await db.refresh(fat)
    assert fat.status_faturamento == "erro"


@pytest.mark.asyncio
async def test_faturamento_bling_nao_enfileira_etiqueta(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """Faturador Bling (correios/agência) NÃO tem fluxo de etiqueta construído →
    o motor não cria comando de etiqueta ao fechar o faturamento."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_dois_faturadores(db, admin)
    await _import_done(client, ["830001", "830002"])

    etq = (await db.execute(
        select(NfCommand).where(NfCommand.action == "imprimir_etiqueta")
    )).scalars().all()
    assert etq == []


@pytest.mark.asyncio
async def test_lease_entrega_comando_etiqueta(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """O comando de etiqueta é entregue no lease com o login do faturador Upseller
    e planilha vazia (a marionete captura o PDF, não sobe arquivo)."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_upseller(db, admin)
    await _ciclo_upseller(client, ["850001"])

    lease = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                              headers={"X-Agent-Token": _TOKEN})
    etq = next(c for c in lease.json()["commands"] if c["action"] == "imprimir_etiqueta")
    assert etq["numeros"] == ["850001"]
    assert etq["ads_power"] == "perfil-U"
    assert etq["usuario"] == "user-u"
    assert etq["senha"] == "segredoU"
    assert etq["planilha_b64"] == ""  # comando de etiqueta não carrega planilha
    # A fila do Upseller lista pelo nº da PLATAFORMA — sem ele a marionete
    # procurava pelo nº do Bling e não achava a linha.
    assert etq["numeros_plataforma"] == ["PLAT-850001"]


@pytest.mark.asyncio
async def test_lease_numero_plataforma_cai_no_proprio_numero(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """Pedido sem numeroloja (avulso digitado à mão) → o nº de busca é o próprio
    número do Bling, mantendo o comportamento antigo."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_dois_faturadores(db, admin)  # _seed_pedido sem numeroloja
    await client.post("/api/nf-cadastro/faturamento/enfileirar",
                      json={"numeros": ["830001"]})

    lease = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                              headers={"X-Agent-Token": _TOKEN})
    cmd = lease.json()["commands"][0]
    assert cmd["numeros"] == ["830001"]
    assert cmd["numeros_plataforma"] == ["830001"]


@pytest.mark.asyncio
async def test_result_etiqueta_marca_status_etiqueta(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """result 'done' num comando de etiqueta marca status_etiqueta='ok' (não toca
    o status_faturamento, que já estava 'ok')."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_upseller(db, admin)
    await _ciclo_upseller(client, ["850001"])
    lease = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                              headers={"X-Agent-Token": _TOKEN})
    etq = next(c for c in lease.json()["commands"] if c["action"] == "imprimir_etiqueta")

    r = await client.post(f"/api/nf-cadastro/agent/commands/{etq['id']}/result",
                          json={"status": "done"}, headers={"X-Agent-Token": _TOKEN})
    assert r.status_code == 200, r.text

    fat = (await db.execute(
        select(NfFaturamento).where(NfFaturamento.pedido_bling == "850001")
    )).scalar_one()
    await db.refresh(fat)
    assert fat.status_etiqueta == "ok"
    assert fat.erro_etiqueta is None
    assert fat.status_faturamento == "ok"


@pytest.mark.asyncio
async def test_result_etiqueta_move_pedido_pra_enviado_etiqueta(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """Etiqueta capturada → o pedido vai pra "Enviado Etiqueta" (83965) no Bling e
    a situação local espelha, que é o que faz ele aparecer no Controle de Estoque
    pra ser impresso."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    fake = _FakeBlingSituacao()

    async def _fake_client(_session):
        return fake

    monkeypatch.setattr(nf_emissao_gerar, "_bling_client_opt", _fake_client)
    auth_as(admin)
    await _seed_upseller(db, admin)
    await _ciclo_upseller(client, ["850001"])
    lease = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                              headers={"X-Agent-Token": _TOKEN})
    etq = next(c for c in lease.json()["commands"] if c["action"] == "imprimir_etiqueta")

    r = await client.post(f"/api/nf-cadastro/agent/commands/{etq['id']}/result",
                          json={"status": "done"}, headers={"X-Agent-Token": _TOKEN})
    assert r.status_code == 200, r.text

    assert fake.chamadas == [(700001, 83965)]
    sit = (await db.execute(
        select(BlingOrder.situacao).where(BlingOrder.numero == "850001")
    )).scalars().all()
    assert set(sit) == {"83965"}


@pytest.mark.asyncio
async def test_result_etiqueta_falhou_nao_move_situacao(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """Sem etiqueta na mão não há o que imprimir → a situação do Bling fica onde
    está (o pedido não entra na fila do Controle de Estoque)."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    fake = _FakeBlingSituacao()

    async def _fake_client(_session):
        return fake

    monkeypatch.setattr(nf_emissao_gerar, "_bling_client_opt", _fake_client)
    auth_as(admin)
    await _seed_upseller(db, admin)
    await _ciclo_upseller(client, ["850001"])
    lease = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                              headers={"X-Agent-Token": _TOKEN})
    etq = next(c for c in lease.json()["commands"] if c["action"] == "imprimir_etiqueta")

    r = await client.post(f"/api/nf-cadastro/agent/commands/{etq['id']}/result",
                          json={"status": "failed", "result": "sem etiqueta"},
                          headers={"X-Agent-Token": _TOKEN})
    assert r.status_code == 200, r.text

    assert fake.chamadas == []
    sit = (await db.execute(
        select(BlingOrder.situacao).where(BlingOrder.numero == "850001")
    )).scalars().all()
    assert set(sit) == {"15"}


@pytest.mark.asyncio
async def test_dedupe_nao_reenfileira_etiqueta(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """Um 2º faturamento 'ok' do mesmo pedido NÃO cria um 2º comando de etiqueta
    enquanto já existe um ativo (pending/claimed) na fila."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_upseller(db, admin)
    await _ciclo_upseller(client, ["850001"])
    await _ciclo_upseller(client, ["850001"])  # segunda passada

    etq = (await db.execute(
        select(NfCommand).where(NfCommand.action == "imprimir_etiqueta")
    )).scalars().all()
    assert len(etq) == 1


@pytest.mark.asyncio
async def test_agent_etiqueta_sem_token_401(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    r = await client.post(
        "/api/nf-cadastro/agent/etiqueta",
        data={"pedido_bling": "870002"},
        files={"file": ("etq.pdf", _etiqueta_pdf(), "application/pdf")},
    )
    assert r.status_code == 401


@pytest.mark.asyncio
async def test_agent_etiqueta_pdf_invalido_422(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    r = await client.post(
        "/api/nf-cadastro/agent/etiqueta",
        data={"pedido_bling": "870003"},
        files={"file": ("x.pdf", b"nao sou pdf", "application/pdf")},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 422
    assert r.json()["detail"]["code"] == "nf_etiqueta_invalida"


async def _seed_ml_etiqueta(db: AsyncSession, admin: User, *, numero: str = "860001") -> None:
    """Loja ML: faturador Bling (NF sai do Bling no faturamento) + cadastro de
    Etiqueta 'upseller' (a etiqueta é puxada pelo Upseller, perfil AdsPower 58)."""
    fat = NfFaturador(
        nome="bling avulso ml", modo="bling", nf_cheia=True,
        sku_fonte="principal", nome_fonte="produto", ncm="4202.12.10",
        ads_power="perfil-A", usuario="user-a", senha_enc=encrypt("segredoA"),
    )
    etq = NfEtiqueta(plataforma="ml", modo="upseller", ads_power="58")
    db.add_all([fat, etq])
    await db.flush()
    db.add(StoreInfo(user_id=admin.id, platform="ml", account_name="lML",
                     bling_store_id="960001", nf_faturador_id=fat.id, nf_etiqueta_id=etq.id))
    await db.flush()
    await _seed_pedido(db, numero=numero, loja="960001", sku="dg053.ci", nome="Capa", unit=500)
    await db.commit()


@pytest.mark.asyncio
async def test_import_avulsa_ml_enfileira_import_etiqueta(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """Loja ML (faturador Bling + etiqueta Upseller): ao fechar o faturamento 'ok'
    o motor auto-enfileira UM comando `import_etiqueta` (.xlsx NF-e=NÃO, no perfil
    AdsPower do cadastro de Etiqueta = 58, sem faturador_id), NÃO um
    `imprimir_etiqueta` ainda (esse vem depois que a importação fecha)."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_ml_etiqueta(db, admin)
    await _import_done(client, ["860001"])

    imp = (await db.execute(
        select(NfCommand).where(NfCommand.action == "import_etiqueta")
    )).scalars().all()
    assert len(imp) == 1
    cmd = imp[0]
    assert cmd.numeros == ["860001"]
    assert cmd.status == "pending"
    assert cmd.source == "auto"
    assert cmd.ads_power == "58"
    assert cmd.faturador_id is None
    assert cmd.nome_arquivo.endswith(".xlsx")
    assert cmd.planilha[:4] == b"PK\x03\x04"

    # NF-e=Não (a NF já saiu do Bling; o Upseller entra só pra puxar a etiqueta).
    # A loja é a CONTA de marketplace (store_info.account_name), registrada como
    # Loja no Upseller — não a genérica "Loja Padrão".
    ws = load_workbook(io.BytesIO(cmd.planilha)).active
    from app.services.nf_upseller import _HEADERS
    col = _HEADERS.index("Necessita Emitir NF-e*") + 1
    assert ws.cell(row=4, column=col).value == "Não"
    col_loja = _HEADERS.index("Nome da Loja*") + 1
    assert ws.cell(row=4, column=col_loja).value == "lML"

    # ainda NÃO existe imprimir_etiqueta (só nasce quando o import_etiqueta fecha).
    pri = (await db.execute(
        select(NfCommand).where(NfCommand.action == "imprimir_etiqueta")
    )).scalars().all()
    assert pri == []

    fat = (await db.execute(
        select(NfFaturamento).where(NfFaturamento.pedido_bling == "860001")
    )).scalar_one()
    await db.refresh(fat)
    assert fat.status_faturamento == "ok"
    assert fat.status_etiqueta == "processando"


@pytest.mark.asyncio
async def test_import_etiqueta_done_encadeia_imprimir_etiqueta(
    db: AsyncSession, client: AsyncClient, admin: User,
    auth_as: Callable[[User | None], None], monkeypatch: pytest.MonkeyPatch,
):
    """`import_etiqueta` 'done' encadeia o `imprimir_etiqueta` no MESMO perfil
    AdsPower (58) do cadastro de Etiqueta — a marionete captura o PDF depois."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    auth_as(admin)
    await _seed_ml_etiqueta(db, admin)
    await _import_done(client, ["860001"])

    lease = await client.post("/api/nf-cadastro/agent/lease", json={"limit": 5},
                              headers={"X-Agent-Token": _TOKEN})
    imp = next(c for c in lease.json()["commands"] if c["action"] == "import_etiqueta")
    assert imp["ads_power"] == "58"

    r = await client.post(f"/api/nf-cadastro/agent/commands/{imp['id']}/result",
                          json={"status": "done"}, headers={"X-Agent-Token": _TOKEN})
    assert r.status_code == 200, r.text

    pri = (await db.execute(
        select(NfCommand).where(NfCommand.action == "imprimir_etiqueta")
    )).scalars().all()
    assert len(pri) == 1
    assert pri[0].numeros == ["860001"]
    assert pri[0].ads_power == "58"
    assert pri[0].status == "pending"
    assert pri[0].source == "auto"


def _lote_pdf(paginas: list[list[str]]) -> bytes:
    """PDF do LOTE sintético: cada item é uma página com as linhas dadas."""
    import fitz

    doc = fitz.open()
    for linhas in paginas:
        page = doc.new_page(width=300, height=442)
        y = 20
        for linha in linhas:
            page.insert_text((20, y), linha, fontsize=8, fontname="helv")
            y += 14
    return doc.tobytes()


@pytest.mark.asyncio
async def test_agent_etiqueta_lote_casa_grava_e_processa(
    db: AsyncSession, client: AsyncClient, admin: User, monkeypatch: pytest.MonkeyPatch
):
    """PDF único do lote → fatia por pedido, casa (numeroloja na Shopee; nome+SKU
    na TikTok), transforma+grava, marca status_etiqueta='ok', fecha o comando
    `imprimir_etiqueta` coberto e move os pedidos pra 83965 no Bling. A fatia
    sem dono volta em `nao_casadas` (nunca chuta)."""
    import fitz

    from app.models import NfEtiquetaArquivo

    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    fake = _FakeBlingSituacao()

    async def _fake_client(_session):
        return fake

    monkeypatch.setattr(nf_emissao_gerar, "_bling_client_opt", _fake_client)

    # Shopee: casa pelo nº da plataforma impresso na etiqueta (numeroloja).
    await _seed_pedido(
        db, numero="890001", loja="930001", sku="dg053.ci", nome="Capa",
        unit=500, bling_id=920001, numeroloja="260808LOTE01",
    )
    # TikTok: sem nº na etiqueta → casa pelo destinatário + SKU da declaração
    # (precisa estar na janela recente do casamento reserva).
    db.add(BlingOrder(
        bling_id=920002, numero="890002", data=datetime.now(UTC),
        loja="930002", situacao="15", item_index=0, item_codigo="b011.20",
        item_descricao="Mala Lisa", item_quantidade=1, itemvalor=161,
        nome_destinatario="Maria Aparecida Silva",
    ))
    # Comando de etiqueta individual coberto pelo lote → fecha 'done'.
    db.add(NfCommand(
        action="imprimir_etiqueta", numeros=["890001", "890002"],
        planilha=b"", nome_arquivo="", status="pending", source="auto",
    ))
    # Coberto só EM PARTE (999999 não veio no lote) → segue pending.
    db.add(NfCommand(
        action="imprimir_etiqueta", numeros=["890001", "999999"],
        planilha=b"", nome_arquivo="", status="pending", source="auto",
    ))
    await db.commit()

    pdf = _lote_pdf([
        [
            "DESTINATÁRIO", "Cleso Menezes", "Pedido: 260808LOTE01",
            "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
            "REMETENTE", "Loja Origem XYZ",
        ],
        ["DECLARAÇÃO DE CONTEÚDO", "SKU", "dg053.ci"],
        ["DESTINATÁRIO", "Maria Aparecida Silva"],
        ["DECLARAÇÃO DE CONTEÚDO", "SKU", "b011.20"],
        ["DESTINATÁRIO", "Cliente Desconhecido"],  # órfã → nao_casadas
    ])
    r = await client.post(
        "/api/nf-cadastro/agent/etiqueta-lote",
        files={"file": ("lote.pdf", pdf, "application/pdf")},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total_fatias"] == 3
    assert body["casadas"] == ["890001", "890002"]
    assert body["falhas"] == []
    assert len(body["nao_casadas"]) == 1
    assert body["nao_casadas"][0]["fatia"] == 2
    assert body["comandos_resolvidos"] == 1

    db.expire_all()
    # etiquetas gravadas TRANSFORMADAS (remetente=destinatário, fatia inteira)
    etq1 = (await db.execute(
        select(NfEtiquetaArquivo).where(NfEtiquetaArquivo.pedido_bling == "890001")
    )).scalar_one()
    doc1 = fitz.open(stream=etq1.blob, filetype="pdf")
    assert doc1.page_count == 2  # etiqueta + declaração juntas
    assert "Loja Origem" not in doc1[0].get_text()
    etq2 = (await db.execute(
        select(NfEtiquetaArquivo).where(NfEtiquetaArquivo.pedido_bling == "890002")
    )).scalar_one()
    assert etq2.size_bytes == len(etq2.blob)

    # status_etiqueta='ok' nos casados
    fats = (await db.execute(
        select(NfFaturamento).where(
            NfFaturamento.pedido_bling.in_(["890001", "890002"])
        )
    )).scalars().all()
    assert {f.pedido_bling: f.status_etiqueta for f in fats} == {
        "890001": "ok", "890002": "ok",
    }

    # comando coberto fechou; o parcialmente coberto segue na fila
    cmds = (await db.execute(
        select(NfCommand).where(NfCommand.action == "imprimir_etiqueta")
    )).scalars().all()
    por_numeros = {tuple(c.numeros): c for c in cmds}
    fechado = por_numeros[("890001", "890002")]
    assert fechado.status == "done"
    assert fechado.result == "etiqueta capturada pelo lote"
    assert por_numeros[("890001", "999999")].status == "pending"

    # pedidos movidos pra "Enviado Etiqueta" (83965) no Bling + espelho local
    assert set(fake.chamadas) == {(920001, 83965), (920002, 83965)}
    sit = (await db.execute(
        select(BlingOrder.situacao).where(
            BlingOrder.numero.in_(["890001", "890002"])
        )
    )).scalars().all()
    assert set(sit) == {"83965"}


@pytest.mark.asyncio
async def test_agent_etiqueta_lote_danfe_numeroloja_solto(
    db: AsyncSession, client: AsyncClient, admin: User, monkeypatch: pytest.MonkeyPatch
):
    """Layout DANFE SIMPLIFICADO (Shopee): o "Pedido:" da etiqueta traz o
    código de triagem (SOCSP7) e o numeroloja real aparece SOLTO no corpo —
    casa mesmo assim, porque TODOS os candidatos são validados no banco."""
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    fake = _FakeBlingSituacao()

    async def _fake_client(_session):
        return fake

    monkeypatch.setattr(nf_emissao_gerar, "_bling_client_opt", _fake_client)
    await _seed_pedido(
        db, numero="890010", loja="930010", sku="uaf001m1.110", nome="AirFryer",
        unit=500, bling_id=920010, numeroloja="260808STCA6RQ5",
    )
    await db.commit()

    pdf = _lote_pdf([
        [
            "Pedido:", "SOCSP7", "DESTINATÁRIO", "Julia Vieira",
            "260808STCA6RQ5", "DANFE SIMPLIFICADO - ETIQUETA",
        ],
        ["DECLARAÇÃO DE CONTEÚDO", "uaf001m1.110"],
    ])
    r = await client.post(
        "/api/nf-cadastro/agent/etiqueta-lote",
        files={"file": ("lote.pdf", pdf, "application/pdf")},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["casadas"] == ["890010"]
    assert body["nao_casadas"] == []
    assert body["falhas"] == []


@pytest.mark.asyncio
async def test_agent_etiqueta_lote_pdf_invalido_422(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    r = await client.post(
        "/api/nf-cadastro/agent/etiqueta-lote",
        files={"file": ("lote.pdf", b"nao sou pdf", "application/pdf")},
        headers={"X-Agent-Token": _TOKEN},
    )
    assert r.status_code == 422
    assert r.json()["detail"]["code"] == "nf_etiqueta_lote_invalido"


@pytest.mark.asyncio
async def test_agent_etiqueta_lote_sem_token_401(
    db: AsyncSession, client: AsyncClient, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(get_settings(), "nf_agent_token", _TOKEN)
    r = await client.post(
        "/api/nf-cadastro/agent/etiqueta-lote",
        files={"file": ("lote.pdf", _lote_pdf([["x"]]), "application/pdf")},
    )
    assert r.status_code == 401
