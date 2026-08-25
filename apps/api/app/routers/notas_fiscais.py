"""Consulta e export de NF-e das contas `bling_notas`.

A tabela `bling_notas` guarda só as credenciais OAuth por conta — as notas
ficam no Bling e são consultadas on-demand via GET /nfe (lista) e
GET /nfe/{id} (detalhe, que traz o link do XML autorizado).

Fluxo:
  * lista: 1 chamada /nfe por página (100/pg) por conta selecionada.
  * exports (XML e XLSX): lista + 1 chamada de detalhe por nota (pra
    resolver o link do XML) + download do arquivo. O XLSX segue o layout
    do "NF-e Report" (28 colunas) e os campos fiscais (CNPJ, série, CFOP,
    ICMS/ST/IPI/PIS/COFINS, dest. UF/CEP) vêm do próprio XML — por isso
    os dois exports pagam o mesmo custo. O detalhe passa pelo rate gate
    global do Bling (5 req/s compartilhado), então ambos são limitados a
    MAX_XML_NOTAS por request pra não estourar o timeout do proxy.

Tokens: o cron `bling_notas_token_refresh` renova a cada 5h (AT dura 6h);
se um token estiver vencido aqui, fazemos o refresh on-demand reusando os
helpers do próprio cron (commit imediato — o Bling rotaciona o RT).
"""
from __future__ import annotations

import asyncio
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Awaitable, Callable
from datetime import UTC, date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Annotated, Any, Literal
from uuid import UUID
from zoneinfo import ZoneInfo

import httpx
import structlog
from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel
from sqlalchemy import func, literal, select
from sqlalchemy.dialects.postgresql import aggregate_order_by
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.db import get_session
from app.deps.auth import require_permission
from app.models import (
    BackgroundJob,
    BackgroundJobStatus,
    BackgroundJobType,
    BlingNota,
    BlingNotaEmitida,
    BlingOrder,
    User,
)
from app.models.enums import UserRole
from app.models.nf import NfEtiquetaArquivo
from app.models.pricing import StoreInfo
from app.schemas.products import JobCreatedOut
from app.services.pos_vendas import (
    JANELA_ANTES_DIAS,
    JANELA_DEPOIS_DIAS,
    Casamento,
    NotaIn,
    PedidoIn,
    match_notas,
)
from app.services.bling_notas_token_refresh import (
    _apply_token_payload,
    _post_oauth_token,
)
from app.services.marketplaces.bling import (
    BLING_API_BASE,
    _acquire_bling_rate_slot,
)
from app.worker_pool import get_arq_pool

logger = structlog.get_logger()
router = APIRouter(prefix="/api/notas-fiscais", tags=["notas_fiscais"])
SAO_PAULO = ZoneInfo("America/Sao_Paulo")

# Bling pagina /nfe em até 100 itens; trava de segurança por conta.
PAGE_SIZE = 100
MAX_PAGES_PER_CONTA = 60  # 6.000 notas por conta por consulta
# Exports = 1 chamada de detalhe por nota a 5 req/s globais — acima
# disso o request passa de ~2min e o proxy derruba com 502.
MAX_XML_NOTAS = 500
MAX_RANGE_DAYS = 184

SITUACOES = {
    1: "Pendente",
    2: "Cancelada",
    3: "Aguardando recibo",
    4: "Rejeitada",
    5: "Autorizada",
    6: "Emitida DANFE",
    7: "Registrada",
    8: "Aguardando protocolo",
    9: "Denegada",
    10: "Consulta situação",
    11: "Bloqueada",
}
TIPOS = {0: "Entrada", 1: "Saída"}


class ContaOut(BaseModel):
    id: UUID
    nome: str


class ContasOut(BaseModel):
    items: list[ContaOut]


class NotaOut(BaseModel):
    conta: str
    bling_id: int
    numero: str | None = None
    data_emissao: str | None = None
    data_operacao: str | None = None
    tipo: str | None = None
    situacao: str | None = None
    cliente: str | None = None
    documento: str | None = None
    valor: float | None = None


class NotasPage(BaseModel):
    items: list[NotaOut]
    total: int
    erros: list[str]


# ─── Bling HTTP ───────────────────────────────────────────────────────


async def _bling_get(token: str, path: str, params: dict | None = None) -> dict:
    """GET autenticado na API v3 com o rate gate global e retry leve em
    429/503 (mesma família de erro tratada no BlingClient)."""
    delay = 1.0
    last_status = 0
    for attempt in range(3):
        await _acquire_bling_rate_slot()
        async with httpx.AsyncClient(timeout=30.0) as c:
            r = await c.get(
                f"{BLING_API_BASE}{path}",
                headers={
                    "Authorization": f"Bearer {token}",
                    "Accept": "application/json",
                },
                params=params,
            )
        last_status = r.status_code
        if r.status_code in (429, 503) and attempt < 2:
            await asyncio.sleep(delay)
            delay *= 2
            continue
        if r.status_code >= 400:
            raise RuntimeError(f"bling http {r.status_code}: {r.text[:200]}")
        return r.json()
    raise RuntimeError(f"bling http {last_status} (retries esgotados)")


async def _ensure_token(s: AsyncSession, conta: BlingNota) -> str:
    """Access token válido da conta; refresh on-demand se vencido."""
    now = datetime.now(UTC)
    if conta.access_token and (
        conta.token_expires_at is None
        or conta.token_expires_at > now + timedelta(seconds=60)
    ):
        return conta.access_token
    if not conta.refresh_token:
        raise RuntimeError("conta sem refresh_token (reautorizar no Bling)")
    payload = await _post_oauth_token(
        conta.basic_auth_b64,
        {"grant_type": "refresh_token", "refresh_token": conta.refresh_token},
    )
    await _apply_token_payload(s, conta.id, payload, clear_code=False)
    await s.refresh(conta)
    if not conta.access_token:
        raise RuntimeError("refresh não retornou access_token")
    return conta.access_token


async def _list_notas_bling(
    token: str, date_from: date, date_to: date
) -> list[dict]:
    out: list[dict] = []
    pagina = 1
    while pagina <= MAX_PAGES_PER_CONTA:
        payload = await _bling_get(
            token,
            "/nfe",
            {
                "pagina": pagina,
                "limite": PAGE_SIZE,
                "dataEmissaoInicial": f"{date_from.isoformat()} 00:00:00",
                "dataEmissaoFinal": f"{date_to.isoformat()} 23:59:59",
            },
        )
        data = payload.get("data") or []
        out.extend(data)
        if len(data) < PAGE_SIZE:
            return out
        pagina += 1
    raise RuntimeError(
        f"mais de {MAX_PAGES_PER_CONTA * PAGE_SIZE} notas no período — "
        "reduza o intervalo de datas"
    )


# ─── shared fetch ─────────────────────────────────────────────────────


async def _resolve_contas(
    s: AsyncSession, conta_ids: list[UUID]
) -> list[BlingNota]:
    stmt = (
        select(BlingNota)
        .where(BlingNota.status == "active")
        .order_by(BlingNota.nome)
    )
    if conta_ids:
        stmt = stmt.where(BlingNota.id.in_(conta_ids))
    rows = (await s.execute(stmt)).scalars().all()
    if not rows:
        raise HTTPException(422, detail={"code": "nenhuma_conta_selecionada"})
    return list(rows)


def _validate_range(date_from: date, date_to: date) -> None:
    if date_to < date_from:
        raise HTTPException(422, detail={"code": "periodo_invalido"})
    if (date_to - date_from).days > MAX_RANGE_DAYS:
        raise HTTPException(
            422,
            detail={
                "code": "periodo_muito_longo",
                "message": f"período máximo de {MAX_RANGE_DAYS} dias",
            },
        )


async def _fetch_all(
    s: AsyncSession,
    conta_ids: list[UUID],
    date_from: date,
    date_to: date,
) -> tuple[list[dict], list[str]]:
    """Notas de todas as contas selecionadas.

    Retorna (rows, erros) — cada row carrega a conta e o token usado (o
    export XML precisa dele pras chamadas de detalhe). Erro em uma conta
    não derruba as demais; vira mensagem em `erros`.
    """
    rows: list[dict] = []
    erros: list[str] = []
    for conta in await _resolve_contas(s, conta_ids):
        try:
            token = await _ensure_token(s, conta)
            notas = await _list_notas_bling(token, date_from, date_to)
        except Exception as e:  # noqa: BLE001
            erros.append(f"{conta.nome}: {str(e)[:200]}")
            logger.warning(
                "nf_fetch_conta_failed", conta=conta.nome, err=str(e)[:300]
            )
            continue
        rows.extend(
            {"conta": conta.nome, "token": token, "nota": n} for n in notas
        )
    rows.sort(key=lambda r: r["nota"].get("dataEmissao") or "", reverse=True)
    return rows, erros


def _to_out(row: dict) -> NotaOut:
    n = row["nota"]
    contato = n.get("contato") or {}
    valor = n.get("valorNota")
    return NotaOut(
        conta=row["conta"],
        bling_id=int(n.get("id") or 0),
        numero=str(n["numero"]) if n.get("numero") is not None else None,
        data_emissao=n.get("dataEmissao"),
        data_operacao=n.get("dataOperacao"),
        tipo=TIPOS.get(n.get("tipo"), str(n.get("tipo") or "")) or None,
        situacao=SITUACOES.get(n.get("situacao"), str(n.get("situacao") or ""))
        or None,
        cliente=contato.get("nome"),
        documento=contato.get("numeroDocumento"),
        valor=float(valor) if valor is not None else None,
    )


_PERM = require_permission("notas_fiscais", "view")
_ContaIds = Annotated[list[UUID], Query(alias="conta")]


# ─── endpoints ────────────────────────────────────────────────────────


@router.get("/contas", response_model=ContasOut)
async def list_contas(
    session: Annotated[AsyncSession, Depends(get_session)],
    _u: Annotated[User, Depends(_PERM)],
) -> ContasOut:
    rows = (
        await session.execute(
            select(BlingNota)
            .where(BlingNota.status == "active")
            .order_by(BlingNota.nome)
        )
    ).scalars().all()
    return ContasOut(items=[ContaOut(id=r.id, nome=r.nome) for r in rows])


@router.get("", response_model=NotasPage)
async def list_notas(
    session: Annotated[AsyncSession, Depends(get_session)],
    _u: Annotated[User, Depends(_PERM)],
    date_from: date = Query(...),
    date_to: date = Query(...),
    conta: _ContaIds = [],  # noqa: B006
) -> NotasPage:
    _validate_range(date_from, date_to)
    rows, erros = await _fetch_all(session, conta, date_from, date_to)
    return NotasPage(
        items=[_to_out(r) for r in rows], total=len(rows), erros=erros
    )


async def _download_xml(url: str) -> bytes | None:
    """XML fica num link assinado (S3) — fora da API, sem rate gate."""
    try:
        async with httpx.AsyncClient(timeout=30.0) as c:
            r = await c.get(url)
        if r.status_code >= 400:
            return None
        return r.content
    except httpx.HTTPError:
        return None


async def _fetch_nota_xml(
    token: str, nota: dict
) -> tuple[dict, bytes | None, str | None]:
    """(detalhe, xml_bytes, motivo_falha) de uma nota.

    `motivo_falha` preenchido quando o XML não está disponível (nota não
    autorizada, link vencido, erro HTTP) — os exports transformam isso em
    aviso sem derrubar o restante do lote.
    """
    try:
        payload = await _bling_get(token, f"/nfe/{nota['id']}")
    except Exception as e:  # noqa: BLE001
        return {}, None, f"erro no detalhe ({str(e)[:120]})"
    detail: dict[str, Any] = payload.get("data") or {}
    xml_url = detail.get("xml")
    if not xml_url:
        situ = SITUACOES.get(nota.get("situacao"), "?")
        return detail, None, f"sem XML disponível (situação {situ})"
    content = await _download_xml(xml_url)
    if content is None:
        return detail, None, "falha ao baixar o XML"
    return detail, content, None


def _xml_filename(conta: str, detail: dict, nota: dict) -> str:
    chave = detail.get("chaveAcesso") or ""
    if chave:
        base = chave
    else:
        base = f"nfe_{nota.get('numero') or nota.get('id')}"
    safe_conta = conta.replace("/", "_")
    return f"{safe_conta}/{base}.xml"


def _check_export_size(rows: list[dict], erros: list[str]) -> None:
    if not rows and erros:
        raise HTTPException(502, detail={"code": "bling_falhou", "erros": erros})
    if len(rows) > MAX_XML_NOTAS:
        raise HTTPException(
            422,
            detail={
                "code": "muitas_notas",
                "message": (
                    f"{len(rows)} notas no período — o export é limitado a "
                    f"{MAX_XML_NOTAS} por vez, reduza o período ou as contas"
                ),
            },
        )


# ─── parse do XML da NF-e (campos do layout "NF-e Report") ────────────

_OPERACAO = {"0": "Inbound", "1": "Outbound"}
_FIN_NFE = {"1": "normal", "2": "complementar", "3": "ajuste", "4": "devolucao"}
# Situações Bling com NF-e efetivamente emitida.
_SITUACOES_ISSUED = {5, 6, 7}


def _first_el(scope: ET.Element | None, name: str) -> ET.Element | None:
    """Primeiro elemento com esse localname (ignora namespace), em qualquer
    profundidade — mesma abordagem namespace-agnóstica do nf_upload."""
    if scope is None:
        return None
    for el in scope.iter():
        if el.tag.rsplit("}", 1)[-1] == name:
            return el
    return None


def _el_text(scope: ET.Element | None, name: str) -> str:
    el = _first_el(scope, name)
    return (el.text or "").strip() if el is not None else ""


def _el_money(scope: ET.Element | None, name: str) -> float:
    try:
        return float(_el_text(scope, name) or 0)
    except ValueError:
        return 0.0


def _parse_nfe_xml(content: bytes) -> dict | None:
    """Campos fiscais do XML autorizado (emit, dest, ide, ICMSTot)."""
    try:
        root = ET.fromstring(content)
    except ET.ParseError:
        return None
    ide = _first_el(root, "ide")
    emit = _first_el(root, "emit")
    dest = _first_el(root, "dest")
    tot = _first_el(root, "ICMSTot")
    det = _first_el(root, "det")
    chave = _el_text(root, "chNFe")
    if not chave:
        inf = _first_el(root, "infNFe")
        if inf is not None:
            chave = (inf.get("Id") or "").removeprefix("NFe")
    dh = _el_text(ide, "dhEmi") or _el_text(ide, "dEmi")
    return {
        "conta_nf": _el_text(emit, "xNome"),
        "cnpj": _el_text(emit, "CNPJ"),
        "serie": _el_text(ide, "serie"),
        "numero": _el_text(ide, "nNF"),
        "operacao": _OPERACAO.get(_el_text(ide, "tpNF"), ""),
        "chave": chave,
        "tipo": _FIN_NFE.get(_el_text(ide, "finNFe"), "normal"),
        "cfop": _el_text(det, "CFOP"),
        "cliente": _el_text(dest, "xNome"),
        "uf_dest": _el_text(dest, "UF"),
        "cep": _el_text(dest, "CEP"),
        "valor_nf": _el_money(tot, "vNF"),
        "base_icms": _el_money(tot, "vBC"),
        "vlr_icms": _el_money(tot, "vICMS"),
        "base_st": _el_money(tot, "vBCST"),
        "vlr_st": _el_money(tot, "vST"),
        "vlr_fcp_st": _el_money(tot, "vFCPST"),
        "vlr_ipi": _el_money(tot, "vIPI"),
        "vlr_ipi_dev": _el_money(tot, "vIPIDevol"),
        "vlr_pis": _el_money(tot, "vPIS"),
        "vlr_cofins": _el_money(tot, "vCOFINS"),
        "vlr_seguro": _el_money(tot, "vSeg"),
        "vlr_outros": _el_money(tot, "vOutro"),
        "vlr_desconto": _el_money(tot, "vDesc"),
        "vlr_frete": _el_money(tot, "vFrete"),
        "data": dh[:19].replace("T", " "),
    }


def _estado_nota(situacao: int | None) -> str:
    if situacao in _SITUACOES_ISSUED:
        return "Issued"
    if situacao == 2:
        return "Cancelled"
    return SITUACOES.get(situacao, str(situacao or ""))


_XLSX_HEADERS = [
    "Conta da Nota Fiscal",
    "CNPJ",
    "Nº de Série",
    "Nº da NF-e",
    "Operação",
    "Chave",
    "Tipo",
    "CFOP",
    "Cliente",
    "Destinatário",
    "CEP",
    "Valor da Nota Fiscal",
    "Base ICMS",
    "Vlr.ICMS",
    "Base ST",
    "Vlr.ST",
    "Vlr.FCP-ST",
    "Valr.IPI",
    "Vlr.IPIDev.",
    "Vlr.PIS",
    "Vlr.COFINS",
    "Vlr.Seguro",
    "Vlr.Outros",
    "Vlr.Desconto",
    "Vlr.Frete",
    "Estado",
    "Data",
    "Observação",
]


XLSX_MEDIA = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
ZIP_MEDIA = "application/zip"
# Teto do export síncrono é MAX_XML_NOTAS (500). O job async sobe o teto
# mas precisa caber no job_timeout do worker (1800s) a ~3-5 req/s no rate
# gate global do Bling — 3000 notas ≈ 10-16 min, com folga.
MAX_EXPORT_NOTAS_ASYNC = 3000

_OnProgress = Callable[[int], Awaitable[None]] | None


def _xlsx_fname(date_from: date, date_to: date) -> str:
    now_sp = datetime.now(SAO_PAULO)
    return (
        f"NF-e_Report_excel_{date_from.strftime('%Y%m%d')}"
        f"_ate_{date_to.strftime('%Y%m%d')}_{now_sp.strftime('%m%d%H%M%S')}.xlsx"
    )


def _xml_zip_fname() -> str:
    return f"notas_fiscais_xml_{datetime.now(SAO_PAULO).strftime('%Y%m%d_%H%M')}.zip"


def _xlsx_row(r: dict, parsed: dict | None, detail: dict, estado: str) -> list:
    """Uma linha (28 colunas) do layout NF-e Report. Sem XML parseado, cai
    pro fallback com o que a listagem tem e colunas fiscais em branco."""
    nota = r["nota"]
    if parsed is None:
        contato = nota.get("contato") or {}
        return [
            r["conta"], "", "", str(nota.get("numero") or ""), "",
            detail.get("chaveAcesso") or "", "", "",
            contato.get("nome") or "", "", "", "", "", "", "", "", "", "",
            "", "", "", "", "", "", "",
            estado, (nota.get("dataEmissao") or "")[:19], "",
        ]
    return [
        parsed["conta_nf"], parsed["cnpj"], parsed["serie"], parsed["numero"],
        parsed["operacao"], parsed["chave"], parsed["tipo"], parsed["cfop"],
        parsed["cliente"], parsed["uf_dest"], parsed["cep"], parsed["valor_nf"],
        parsed["base_icms"], parsed["vlr_icms"], parsed["base_st"],
        parsed["vlr_st"], parsed["vlr_fcp_st"], parsed["vlr_ipi"],
        parsed["vlr_ipi_dev"], parsed["vlr_pis"], parsed["vlr_cofins"],
        parsed["vlr_seguro"], parsed["vlr_outros"], parsed["vlr_desconto"],
        parsed["vlr_frete"], estado, parsed["data"], "",
    ]


async def build_xml_zip(
    rows: list[dict], erros: list[str], on_progress: _OnProgress = None
) -> tuple[bytes, int]:
    """ZIP com os XMLs das notas (uma pasta por conta). Notas sem XML
    disponível entram no `_avisos.txt`. Retorna (bytes, nº de avisos).
    `on_progress(done)` é chamado a cada nota (export async)."""
    avisos: list[str] = list(erros)
    buf = BytesIO()
    seen: set[str] = set()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for i, r in enumerate(rows):
            nota = r["nota"]
            label = f"{r['conta']} nº {nota.get('numero') or nota.get('id')}"
            detail, content, motivo = await _fetch_nota_xml(r["token"], nota)
            if content is None:
                avisos.append(f"{label}: {motivo}")
            else:
                name = _xml_filename(r["conta"], detail, nota)
                if name in seen:
                    name = name.replace(".xml", f"_{nota['id']}.xml")
                seen.add(name)
                zf.writestr(name, content)
            if on_progress is not None:
                await on_progress(i + 1)
        if avisos:
            zf.writestr("_avisos.txt", "\n".join(avisos))
    return buf.getvalue(), len(avisos)


async def build_xlsx(
    rows: list[dict], erros: list[str], on_progress: _OnProgress = None
) -> tuple[bytes, int]:
    """Planilha no layout "NF-e Report" — campos fiscais extraídos do XML de
    cada nota. Nota sem XML entra com os dados da listagem e colunas fiscais
    em branco, mais a aba "Avisos". Retorna (bytes, nº de avisos)."""
    avisos: list[str] = list(erros)
    wb = Workbook()
    ws = wb.active
    ws.title = "NF-e"
    ws.append(_XLSX_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, r in enumerate(rows):
        nota = r["nota"]
        label = f"{r['conta']} nº {nota.get('numero') or nota.get('id')}"
        estado = _estado_nota(nota.get("situacao"))
        detail, content, motivo = await _fetch_nota_xml(r["token"], nota)
        parsed = _parse_nfe_xml(content) if content else None
        if parsed is None:
            avisos.append(
                f"{label}: {motivo or 'XML inválido (não parseou)'}"
            )
        ws.append(_xlsx_row(r, parsed, detail, estado))
        if on_progress is not None:
            await on_progress(i + 1)

    if avisos:
        ws_av = wb.create_sheet("Avisos")
        ws_av.append(["Aviso"])
        ws_av["A1"].font = Font(bold=True)
        for a in avisos:
            ws_av.append([a])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(avisos)


@router.get("/export.xml")
async def export_xml(
    session: Annotated[AsyncSession, Depends(get_session)],
    _u: Annotated[User, Depends(_PERM)],
    date_from: date = Query(...),
    date_to: date = Query(...),
    conta: _ContaIds = [],  # noqa: B006
) -> StreamingResponse:
    """ZIP com os XMLs das notas do período (síncrono, até MAX_XML_NOTAS)."""
    _validate_range(date_from, date_to)
    rows, erros = await _fetch_all(session, conta, date_from, date_to)
    _check_export_size(rows, erros)
    data, n_avisos = await build_xml_zip(rows, erros)
    logger.info(
        "nf_export_xml",
        notas=len(rows),
        avisos=n_avisos,
        date_from=str(date_from),
        date_to=str(date_to),
    )
    return StreamingResponse(
        BytesIO(data),
        media_type=ZIP_MEDIA,
        headers={
            "Content-Disposition": f'attachment; filename="{_xml_zip_fname()}"'
        },
    )


@router.get("/export.xlsx")
async def export_xlsx(
    session: Annotated[AsyncSession, Depends(get_session)],
    _u: Annotated[User, Depends(_PERM)],
    date_from: date = Query(...),
    date_to: date = Query(...),
    conta: _ContaIds = [],  # noqa: B006
) -> StreamingResponse:
    """Planilha "NF-e Report" do período (síncrono, até MAX_XML_NOTAS)."""
    _validate_range(date_from, date_to)
    rows, erros = await _fetch_all(session, conta, date_from, date_to)
    _check_export_size(rows, erros)
    data, n_avisos = await build_xlsx(rows, erros)
    logger.info(
        "nf_export_xlsx",
        notas=len(rows),
        avisos=n_avisos,
        date_from=str(date_from),
        date_to=str(date_to),
    )
    return StreamingResponse(
        BytesIO(data),
        media_type=XLSX_MEDIA,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{_xlsx_fname(date_from, date_to)}"'
            )
        },
    )


# ─── export assíncrono (lotes grandes, via worker) ────────────────────


class ExportJobIn(BaseModel):
    fmt: Literal["xlsx", "xml"] = "xlsx"
    date_from: date
    date_to: date
    conta: list[UUID] = []


@router.post(
    "/export-job",
    response_model=JobCreatedOut,
    status_code=status.HTTP_201_CREATED,
)
async def enqueue_export_job(
    body: ExportJobIn,
    session: Annotated[AsyncSession, Depends(get_session)],
    user: Annotated[User, Depends(_PERM)],
) -> JobCreatedOut:
    """Enfileira o export como BackgroundJob — sem o teto de 500 do
    síncrono (o worker foge do timeout do proxy). O front faz polling em
    GET /api/jobs/{id} e baixa em GET /export-job/{id}/download."""
    _validate_range(body.date_from, body.date_to)
    job = BackgroundJob(
        type=BackgroundJobType.EXPORT_NOTAS_FISCAIS,
        status=BackgroundJobStatus.PENDING,
        created_by=user.id,
        payload={
            "fmt": body.fmt,
            "date_from": body.date_from.isoformat(),
            "date_to": body.date_to.isoformat(),
            "conta": [str(c) for c in body.conta],
        },
    )
    session.add(job)
    await session.flush()

    pool = await get_arq_pool()
    arq = await pool.enqueue_job("export_notas_run", str(job.id))
    if arq is not None:
        job.arq_job_id = arq.job_id
    await session.commit()
    return JobCreatedOut(job_id=job.id)


@router.get("/export-job/{job_id}/download")
async def download_export_job(
    job_id: UUID,
    session: Annotated[AsyncSession, Depends(get_session)],
    user: Annotated[User, Depends(_PERM)],
) -> FileResponse:
    """Baixa o arquivo gerado por um export job concluído."""
    job = await session.get(BackgroundJob, job_id)
    if job is None or job.type != BackgroundJobType.EXPORT_NOTAS_FISCAIS:
        raise HTTPException(404, detail={"code": "job_not_found"})
    if user.role != UserRole.ADMIN and job.created_by != user.id:
        raise HTTPException(404, detail={"code": "job_not_found"})
    if job.status != BackgroundJobStatus.SUCCEEDED:
        raise HTTPException(409, detail={"code": "job_nao_concluido"})
    rel = (job.result or {}).get("file_path")
    abs_path = Path(get_settings().uploads_dir) / rel if rel else None
    if not rel or abs_path is None or not abs_path.exists():
        raise HTTPException(404, detail={"code": "arquivo_indisponivel"})
    return FileResponse(
        abs_path,
        media_type=job.result.get("media_type") or "application/octet-stream",
        filename=job.result.get("filename") or abs_path.name,
    )


# ─── Pós Vendas ───────────────────────────────────────────────────────
#
# A página /notas-fiscais virou "Pós Vendas": pedidos ENVIADOS no período
# com as duas notas de cada envio (embalagem × produto). Nada de Bling ao
# vivo na listagem — o cron `pos_vendas_notas_sync` mantém o espelho
# `bling_notas_emitidas`; só o download do XML consulta o Bling (o link S3
# expira, regenerar custa 2 chamadas).

MAX_POS_VENDAS_DAYS = 62


class PosVendaNfOut(BaseModel):
    # id de bling_notas_emitidas — vai no GET /pos-vendas/nota/{id}/xml.
    nota_id: UUID
    emitente: str | None = None
    cnpj: str | None = None
    numero: str | None = None
    valor: float | None = None
    data_emissao: str | None = None
    # Como a nota foi casada: "pedido" (complemento == numeroloja, exata)
    # ou "cpf" (CPF + janela de dias) — o front sinaliza a heurística.
    via: str | None = None


class PosVendaRowOut(BaseModel):
    pedido_bling: str
    pedido_marketplace: str | None = None
    # ISO. Com horário quando existe etiqueta gerada (created_at do blob);
    # senão só a data do envio (em_andamento_data).
    data_envio: str | None = None
    envio_com_hora: bool = False
    loja: str | None = None
    plataforma: str | None = None
    sku: str | None = None
    produto: str | None = None
    valor: float | None = None
    nf_embalagem: PosVendaNfOut | None = None
    nf_produto: PosVendaNfOut | None = None


class PosVendasPage(BaseModel):
    items: list[PosVendaRowOut]
    total: int


@router.get("/pos-vendas", response_model=PosVendasPage)
async def list_pos_vendas(
    session: Annotated[AsyncSession, Depends(get_session)],
    user: Annotated[User, Depends(_PERM)],
    date_from: date = Query(...),
    date_to: date = Query(...),
) -> PosVendasPage:
    if date_to < date_from:
        raise HTTPException(422, detail={"code": "periodo_invalido"})
    if (date_to - date_from).days > MAX_POS_VENDAS_DAYS:
        raise HTTPException(
            422,
            detail={
                "code": "periodo_muito_longo",
                "message": f"período máximo de {MAX_POS_VENDAS_DAYS} dias",
            },
        )

    bo = BlingOrder
    pedidos_rows = (
        await session.execute(
            select(
                bo.numero,
                func.max(bo.numeroloja).label("numeroloja"),
                func.max(bo.em_andamento_data).label("envio"),
                func.max(bo.total).label("total"),
                func.max(bo.documento_destinatario).label("doc"),
                func.max(bo.loja).label("loja"),
                func.string_agg(
                    bo.item_codigo,
                    aggregate_order_by(literal(", "), bo.item_index),
                ).label("skus"),
                func.string_agg(
                    bo.item_descricao,
                    aggregate_order_by(literal(" | "), bo.item_index),
                ).label("produtos"),
            )
            .where(
                bo.em_andamento_data >= date_from,
                bo.em_andamento_data <= date_to,
                bo.numero.is_not(None),
            )
            .group_by(bo.numero)
        )
    ).all()

    numeros = [r.numero for r in pedidos_rows]
    etiquetas: dict[str, datetime] = {}
    if numeros:
        for ped, created in (
            await session.execute(
                select(
                    NfEtiquetaArquivo.pedido_bling, NfEtiquetaArquivo.created_at
                ).where(NfEtiquetaArquivo.pedido_bling.in_(numeros))
            )
        ).all():
            etiquetas[ped] = created

    lojas = {r.loja for r in pedidos_rows if r.loja}
    stores: dict[str, StoreInfo] = {}
    if lojas:
        for si in (
            (
                await session.execute(
                    select(StoreInfo).where(StoreInfo.bling_store_id.in_(lojas))
                )
            )
            .scalars()
            .all()
        ):
            if si.bling_store_id:
                stores[si.bling_store_id] = si

    j1 = datetime.combine(
        date_from - timedelta(days=JANELA_ANTES_DIAS), datetime.min.time()
    )
    j2 = datetime.combine(
        date_to + timedelta(days=JANELA_DEPOIS_DIAS), datetime.max.time()
    )
    notas_rows = (
        await session.execute(
            select(BlingNotaEmitida, BlingNota)
            .join(BlingNota, BlingNota.id == BlingNotaEmitida.conta_id)
            .where(
                BlingNotaEmitida.data_emissao.between(j1, j2)
                | BlingNotaEmitida.data_emissao.is_(None)
            )
        )
    ).all()

    notas_in: list[NotaIn] = []
    nota_by_key: dict[Any, tuple[BlingNotaEmitida, BlingNota]] = {}
    for ne, conta in notas_rows:
        notas_in.append(
            NotaIn(
                key=ne.id,
                conta_cnpj=conta.cnpj,
                cpf=ne.cpf_dest,
                complemento=ne.complemento,
                data_emissao=ne.data_emissao,
                situacao=ne.situacao,
            )
        )
        nota_by_key[ne.id] = (ne, conta)

    pedidos_in = [
        PedidoIn(
            numero=r.numero,
            numeroloja=r.numeroloja,
            cpf=r.doc,
            envio=r.envio,
            store_cnpj=(stores[r.loja].cnpj if r.loja in stores else None),
        )
        for r in pedidos_rows
    ]
    casamentos = match_notas(pedidos_in, notas_in)

    def _nf_out(key: Any, via: str | None) -> PosVendaNfOut | None:
        if key is None:
            return None
        ne, conta = nota_by_key[key]
        return PosVendaNfOut(
            nota_id=ne.id,
            emitente=conta.emitente or conta.nome,
            cnpj=conta.cnpj,
            numero=ne.numero,
            valor=float(ne.valor) if ne.valor is not None else None,
            data_emissao=(
                ne.data_emissao.isoformat(sep=" ") if ne.data_emissao else None
            ),
            via=via,
        )

    items: list[PosVendaRowOut] = []
    for r in pedidos_rows:
        c = casamentos.get(r.numero) or Casamento()
        etq = etiquetas.get(r.numero)
        si = stores.get(r.loja) if r.loja else None
        items.append(
            PosVendaRowOut(
                pedido_bling=r.numero,
                pedido_marketplace=r.numeroloja,
                data_envio=(
                    etq.isoformat()
                    if etq
                    else (r.envio.isoformat() if r.envio else None)
                ),
                envio_com_hora=etq is not None,
                loja=si.account_name if si else None,
                plataforma=si.platform if si else None,
                sku=r.skus,
                produto=r.produtos,
                valor=float(r.total) if r.total is not None else None,
                nf_embalagem=_nf_out(c.embalagem, c.embalagem_via),
                nf_produto=_nf_out(c.produto, c.produto_via),
            )
        )
    items.sort(key=lambda i: i.data_envio or "", reverse=True)
    return PosVendasPage(items=items, total=len(items))


@router.get("/pos-vendas/nota/{nota_id}/xml")
async def download_pos_venda_xml(
    nota_id: UUID,
    session: Annotated[AsyncSession, Depends(get_session)],
    user: Annotated[User, Depends(_PERM)],
) -> Response:
    """XML autorizado de UMA nota (botão XML da tabela Pós Vendas)."""
    ne = await session.get(BlingNotaEmitida, nota_id)
    if ne is None:
        raise HTTPException(404, detail={"code": "nota_nao_encontrada"})
    conta = await session.get(BlingNota, ne.conta_id)
    if conta is None:
        raise HTTPException(404, detail={"code": "conta_nao_encontrada"})
    token = await _ensure_token(session, conta)
    _, xml_bytes, motivo = await _fetch_nota_xml(
        token, {"id": ne.bling_id, "situacao": ne.situacao}
    )
    if xml_bytes is None:
        raise HTTPException(
            502,
            detail={
                "code": "xml_indisponivel",
                "message": motivo or "XML indisponível",
            },
        )
    fname = (ne.chave_acesso or f"nfe_{ne.numero or ne.bling_id}") + ".xml"
    return Response(
        content=xml_bytes,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
