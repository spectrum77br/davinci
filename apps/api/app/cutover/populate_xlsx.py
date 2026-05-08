"""Populate companies/stores/cadastros from `__atualizações app.xlsx`.

Reads sheets `empresas` and `cadastro`. Idempotent by:
- companies.cnpj  (UPSERT)
- stores (company_id, marketplace) (UPSERT)
- cadastros (tipo, codigo) (UPSERT)

Cadastros are inserted WITHOUT cadastros_stores links — the marketplace cell
values reference store names that need human mapping. They're stored as JSON
in `cadastros.obs` so the frontend can later present a mapping UI.
"""
# ruff: noqa: S608

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import asyncpg
import openpyxl

from app.config import get_settings

log = logging.getLogger("populate_xlsx")

MARKETPLACES = ("ml", "shopee", "amazon", "aliexpress", "temu", "tiktok", "shein", "magalu", "site")

# How a cell value in a marketplace column maps to store_status.
STATUS_BY_CELL: dict[str, str] = {
    "x": "active",
    "s": "active",
    "x pr": "active",
    "banido": "banned",
    "fechar": "closing",
    "?": "under_review",
}


@dataclass
class Stats:
    companies_inserted: int = 0
    companies_skipped: int = 0
    stores_inserted: int = 0
    cadastros_inserted: int = 0
    unmapped_cell_values: dict[str, int] = field(default_factory=dict)


# --------------------------------------------------------------------------
# Normalization helpers
# --------------------------------------------------------------------------


def _digits(v: Any) -> str | None:
    if v is None:
        return None
    s = "".join(ch for ch in str(v) if ch.isdigit())
    return s or None


def _txt(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    return s or None


def _slug_from_razao(razao: str) -> str:
    # First non-trivial alphabetic word, lowercased.
    s = re.sub(r"\(.*?\)", " ", razao)
    s = re.sub(r"[^A-Za-zÀ-ú0-9 ]", " ", s)
    parts = [p for p in s.lower().split() if p and p not in ("ltda", "s.a.", "sa", "spe")]
    return parts[0] if parts else razao.strip().lower()[:32]


def _store_status_for_cell(cell: Any) -> tuple[str, str | None]:
    """Return (status, notes_extra)."""
    if cell is None:
        return ("", None)
    s = str(cell).replace("\xa0", " ").strip().lower()
    if s in STATUS_BY_CELL:
        return (STATUS_BY_CELL[s], None)
    if not s:
        return ("", None)
    # Anything unrecognized → assume active, keep raw text as notes.
    return ("active", str(cell).strip())


# --------------------------------------------------------------------------
# Reading
# --------------------------------------------------------------------------


def read_sheets(xlsx_path: Path) -> tuple[list[dict], list[dict]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    empresas_rows: list[dict] = []
    ws = wb["empresas"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] is None:
            continue
        razao = _txt(row[0])
        if not razao:
            continue
        empresas_rows.append({
            "razao_social": razao,
            "uf": _txt(row[2]),
            "cnpj": _digits(row[3]),
            "inscricao_estadual": _txt(row[4]),
            "apelido_raw": _txt(row[5]),
            "stores": {mp: row[6 + idx] for idx, mp in enumerate(MARKETPLACES)},
            "obs": _txt(row[15]),
        })

    cadastro_rows: list[dict] = []
    ws = wb["cadastro"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        tipo = _txt(row[0])
        if not tipo:
            continue
        codigo = _txt(row[3])
        if not codigo:
            continue
        cadastro_rows.append({
            "tipo": tipo.lower(),
            "provedor": _txt(row[1]),
            "responsavel_txt": _txt(row[2]),
            "codigo": codigo,
            "raw_links": {
                mp: _txt(row[4 + idx]) for idx, mp in enumerate(MARKETPLACES)
            },
        })

    return empresas_rows, cadastro_rows


# --------------------------------------------------------------------------
# Writing
# --------------------------------------------------------------------------


async def reset_data(conn: asyncpg.Connection, schema: str) -> None:
    log.warning("deleting existing dummy/seed data from %s", schema)
    await conn.execute(f"DELETE FROM {schema}.cadastros_stores")
    await conn.execute(f"DELETE FROM {schema}.cadastros")
    await conn.execute(f"DELETE FROM {schema}.stores")
    await conn.execute(f"DELETE FROM {schema}.companies")


async def upsert_company(conn: asyncpg.Connection, schema: str, e: dict, used_apelidos: set[str]) -> str | None:
    razao = e["razao_social"]
    apelido = e["apelido_raw"] or _slug_from_razao(razao)
    apelido = apelido.strip()

    # Apelido isn't unique in schema, but to keep things tidy we suffix collisions.
    base = apelido
    suffix = 2
    while apelido in used_apelidos:
        apelido = f"{base}-{suffix}"
        suffix += 1
    used_apelidos.add(apelido)

    cnpj = e["cnpj"]
    ie = e["inscricao_estadual"]
    if ie and ie.strip().lower() == "na":
        ie = None

    # ON CONFLICT requires unique key — cnpj is unique. Without cnpj, no upsert key.
    if cnpj:
        row = await conn.fetchrow(
            f"""
            INSERT INTO {schema}.companies
              (razao_social, apelido, uf, cnpj, inscricao_estadual, obs)
            VALUES ($1,$2,$3,$4,$5,$6)
            ON CONFLICT (cnpj) DO UPDATE SET
              razao_social = EXCLUDED.razao_social,
              apelido = EXCLUDED.apelido,
              uf = EXCLUDED.uf,
              inscricao_estadual = EXCLUDED.inscricao_estadual,
              obs = EXCLUDED.obs
            RETURNING id
            """,
            razao, apelido, e["uf"], cnpj, ie, e["obs"],
        )
    else:
        # No CNPJ → unconditional insert (no dedup key available).
        row = await conn.fetchrow(
            f"""
            INSERT INTO {schema}.companies
              (razao_social, apelido, uf, cnpj, inscricao_estadual, obs)
            VALUES ($1,$2,$3,NULL,$4,$5)
            RETURNING id
            """,
            razao, apelido, e["uf"], ie, e["obs"],
        )
    return str(row["id"]) if row else None


async def upsert_stores(conn: asyncpg.Connection, schema: str, company_id: str, e: dict, stats: Stats) -> None:
    for mp, cell in e["stores"].items():
        status, notes = _store_status_for_cell(cell)
        if not status:
            continue
        await conn.execute(
            f"""
            INSERT INTO {schema}.stores (company_id, marketplace, status, notes)
            VALUES ($1, $2::{schema}.marketplace, $3::{schema}.store_status, $4)
            ON CONFLICT (company_id, marketplace) DO UPDATE SET
              status = EXCLUDED.status,
              notes = COALESCE(EXCLUDED.notes, {schema}.stores.notes)
            """,
            company_id, mp, status, notes,
        )
        stats.stores_inserted += 1


async def upsert_cadastro(conn: asyncpg.Connection, schema: str, c: dict) -> None:
    tipo = c["tipo"]
    if tipo not in ("fone", "email", "dominio", "servidor"):
        log.warning("skipping cadastro tipo=%s codigo=%s", tipo, c["codigo"])
        return

    raw = {k: v for k, v in c["raw_links"].items() if v}
    obs = f"responsavel: {c['responsavel_txt']}" if c["responsavel_txt"] else None
    raw_json = json.dumps(raw, ensure_ascii=False)

    existing = await conn.fetchval(
        f"SELECT id FROM {schema}.cadastros WHERE tipo = $1::{schema}.cadastro_tipo AND codigo = $2",
        tipo, c["codigo"],
    )
    if existing:
        await conn.execute(
            f"""
            UPDATE {schema}.cadastros
               SET provedor = $2, obs = $3, raw_links = $4::jsonb
             WHERE id = $1
            """,
            existing, c["provedor"], obs, raw_json,
        )
    else:
        await conn.execute(
            f"""
            INSERT INTO {schema}.cadastros
              (tipo, provedor, codigo, obs, raw_links, status)
            VALUES ($1::{schema}.cadastro_tipo, $2, $3, $4, $5::jsonb,
                    'active'::{schema}.cadastro_status)
            """,
            tipo, c["provedor"], c["codigo"], obs, raw_json,
        )


# --------------------------------------------------------------------------
# Orchestrator
# --------------------------------------------------------------------------


async def run(xlsx_path: Path, *, reset: bool, dry_run: bool) -> Stats:
    settings = get_settings()
    schema = settings.database_schema
    raw_url = settings.database_url.replace("+asyncpg", "")

    empresas, cadastros = read_sheets(xlsx_path)
    log.info("loaded %d empresas, %d cadastros from %s", len(empresas), len(cadastros), xlsx_path)

    stats = Stats()
    conn = await asyncpg.connect(raw_url)
    try:
        async with conn.transaction():
            if reset:
                await reset_data(conn, schema)

            used_apelidos: set[str] = set()
            company_ids: list[tuple[dict, str]] = []
            for e in empresas:
                cid = await upsert_company(conn, schema, e, used_apelidos)
                if cid:
                    company_ids.append((e, cid))
                    stats.companies_inserted += 1
                else:
                    stats.companies_skipped += 1

            for e, cid in company_ids:
                await upsert_stores(conn, schema, cid, e, stats)

            for c in cadastros:
                await upsert_cadastro(conn, schema, c)
                stats.cadastros_inserted += 1

            if dry_run:
                log.warning("dry-run: rolling back")
                raise _DryRunRollback()
    except _DryRunRollback:
        pass
    finally:
        await conn.close()
    return stats


class _DryRunRollback(Exception):
    pass


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--reset", action="store_true", help="DELETE existing rows first")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    stats = asyncio.run(run(args.xlsx, reset=args.reset, dry_run=args.dry_run))
    print(json.dumps({
        "companies_inserted": stats.companies_inserted,
        "companies_skipped": stats.companies_skipped,
        "stores_inserted": stats.stores_inserted,
        "cadastros_inserted": stats.cadastros_inserted,
    }, indent=2))


if __name__ == "__main__":
    main()
