"""xlsx parsing helpers for audit (Fase 10).

Sheets are read with openpyxl in `read_only=True` mode so 5k-row planilhas
don't blow memory. Header detection: first non-empty row. Data rows: every
row after header that has at least one non-empty cell.

The expected layout is one column with header containing "sku" (case
insensitive) plus N price columns whose headers are the planilha's account
names. The account name → `pricing_accounts.id` mapping is supplied
separately by the user via the audit setup UI.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


@dataclass(slots=True)
class SheetPreview:
    sheet_name: str
    headers: list[str]
    sku_column: int | None
    rows: list[list[str | None]]
    total_rows: int


class AuditParseError(Exception):
    pass


def list_sheets(path: str | Path) -> list[str]:
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except InvalidFileException as e:
        raise AuditParseError(f"invalid_xlsx: {e}") from e
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _detect_sku_column(headers: list[str]) -> int | None:
    for i, h in enumerate(headers):
        if h.lower() in {"sku", "sku/codigo", "código", "codigo", "cod"}:
            return i
    return None


def preview(path: str | Path, sheet: str, *, max_rows: int = 10) -> SheetPreview:
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except InvalidFileException as e:
        raise AuditParseError(f"invalid_xlsx: {e}") from e
    try:
        if sheet not in wb.sheetnames:
            raise AuditParseError(f"sheet_not_found: {sheet}")
        ws = wb[sheet]

        rows_iter = ws.iter_rows(values_only=True)
        headers: list[str] = []
        for raw in rows_iter:
            if any(c is not None and str(c).strip() != "" for c in raw):
                headers = [_normalize_header(c) for c in raw]
                break

        data: list[list[str | None]] = []
        total = 0
        for raw in rows_iter:
            if not any(c is not None and str(c).strip() != "" for c in raw):
                continue
            total += 1
            if len(data) < max_rows:
                data.append([_normalize_header(c) or None for c in raw])

        return SheetPreview(
            sheet_name=sheet,
            headers=headers,
            sku_column=_detect_sku_column(headers),
            rows=data,
            total_rows=total,
        )
    finally:
        wb.close()


def to_decimal(raw: object) -> Decimal | None:
    """Parse a cell into a positive Decimal price. Accepts comma-decimal
    Brazilian format ("1.234,56") and plain dot ("1234.56")."""
    if raw is None:
        return None
    if isinstance(raw, int | float | Decimal):
        try:
            d = Decimal(str(raw))
        except InvalidOperation:
            return None
        return d if d > 0 else None
    s = str(raw).strip()
    if not s:
        return None
    s = s.replace("R$", "").replace(" ", " ").strip()
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return d if d > 0 else None


def iter_rows(path: str | Path, sheet: str):
    """Generator yielding (row_number, [cell_str]) for every non-empty data
    row after the header. Header itself is yielded once with row_number=0."""
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise AuditParseError(f"sheet_not_found: {sheet}")
        ws = wb[sheet]
        header_yielded = False
        n = 0
        for raw in ws.iter_rows(values_only=True):
            if not header_yielded:
                if any(c is not None and str(c).strip() != "" for c in raw):
                    yield 0, [_normalize_header(c) for c in raw]
                    header_yielded = True
                continue
            if not any(c is not None and str(c).strip() != "" for c in raw):
                continue
            n += 1
            yield n, list(raw)
    finally:
        wb.close()
