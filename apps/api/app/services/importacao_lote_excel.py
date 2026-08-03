"""Excel de CONFERÊNCIA de um LOTE de importação (restock ML25/ML26/…).

Visão resumida pra o humano analisar/conferir um lote: uma linha por
produto DA CATEGORIA do lote — todos os SKUs da tabela, mesmo sem
quantidade nesse lote (pedido do operador: planilha completa pra
planejar, não só o que já foi comprado). Mostra modelo, SKU, custo
unitário, estoque Bling, quantidade (nesse lote) e total (qtd × custo).
Linhas sem quantidade ficam com Quantidade/Total em branco.

Camada de arquivo pura: recebe as linhas já calculadas (modelo/sku/
custo/estoque/quantidade/total) e devolve os bytes do .xlsx. Não lê banco.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

LOTE_EXCEL_MEDIA = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_HEADERS = ("Modelo Bling", "SKU", "Custo Bling", "Estoque Bling", "Quantidade", "Total")


@dataclass(frozen=True)
class LinhaLote:
    modelo: str | None
    sku: str | None
    custo: Decimal
    estoque: int | None
    quantidade: int
    total: Decimal


def montar_xlsx(nome_lote: str, linhas: Iterable[LinhaLote]) -> bytes:
    """Monta o .xlsx de conferência do lote (uma linha por produto)."""
    wb = Workbook()
    ws = wb.active
    ws.title = (nome_lote or "lote")[:31]

    ws.append(list(_HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)

    total_qtd = 0
    total_valor = Decimal("0.00")
    for ln in linhas:
        qtd = int(ln.quantidade or 0)
        custo = ln.custo if isinstance(ln.custo, Decimal) else Decimal(str(ln.custo or 0))
        total = ln.total if isinstance(ln.total, Decimal) else Decimal(str(ln.total or 0))
        total_qtd += qtd
        total_valor += total
        ws.append([
            ln.modelo or "",
            ln.sku or "",
            float(custo),
            ln.estoque if ln.estoque is not None else None,
            # Sem quantidade no lote → célula em branco (não 0), pra
            # ficar visível o que ainda não entrou na compra.
            qtd if qtd > 0 else None,
            float(total) if qtd > 0 else None,
        ])

    # Linha de total no rodapé.
    ws.append(["", "Total", "", "", total_qtd, float(total_valor)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col, width in zip("ABCDEF", (40, 18, 14, 14, 12, 16)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
