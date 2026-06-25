"""Gera a planilha de rentabilidade por pedido (aba Pedidos + Resumo).

Fonte: agregação por pedido sobre `bling_orders` (mesma fórmula da
rentabilidade diária — base/comissão/frete = MAX por pedido, custo = SUM
de preco_custo×quantidade), situações 6/15/83953, recortada por período.
Sem coluna de categoria. Consumido por GET /api/margens/rentabilidade/export.xlsx.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_FONT = "Arial"
_BRL = 'R$ #,##0.00;[Red]-R$ #,##0.00'
_PCT = '0.0%;[Red]-0.0%'
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(name=_FONT, bold=True, color="FFFFFF", size=10)
_TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
_ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
_thin = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _loja_label(loja: str | None, store_map: Mapping[str, str]) -> str:
    key = str(loja or "")
    return store_map.get(key, key or "(sem loja)")


def _header(ws, row: int, cols: Sequence[str], start: int = 1) -> None:
    for j, h in enumerate(cols):
        cell = ws.cell(row, start + j, h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def build_rentabilidade_xlsx(
    rows: Sequence[Mapping],
    store_map: Mapping[str, str],
    inicio: date,
    fim: date,
) -> BytesIO:
    """Monta o .xlsx e devolve um buffer pronto pra StreamingResponse."""
    wb = Workbook()

    # ---------------- Aba Pedidos ----------------
    ws = wb.active
    ws.title = "Pedidos"
    headers = [
        "Data", "Nº Pedido", "Nº Loja", "Loja", "Situação",
        "Base (R$)", "Comissão (R$)", "Frete (R$)", "Custo (R$)",
        "Lucro (R$)", "Margem %",
    ]
    ws.append(headers)
    _header(ws, 1, headers)

    for r in rows:
        data_sp = r["data_sp"]
        ws.append([
            data_sp.isoformat() if hasattr(data_sp, "isoformat") else str(data_sp or ""),
            r["numero"],
            r["numeroloja"] or "",
            _loja_label(r["loja"], store_map),
            r["situacao_nome"] or "",
            float(r["base"] or 0),
            float(r["com"] or 0),
            float(r["frete"] or 0),
            float(r["custo"] or 0),
            None,  # Lucro (fórmula)
            None,  # Margem (fórmula)
        ])

    n = len(rows)
    last = n + 1  # última linha de dados
    for i in range(2, last + 1):
        ws.cell(i, 10).value = f"=F{i}-G{i}-H{i}-I{i}"        # Lucro
        ws.cell(i, 11).value = f'=IFERROR(J{i}/I{i},"-")'      # Margem = Lucro/Custo
        for c in (6, 7, 8, 9, 10):
            ws.cell(i, c).number_format = _BRL
        ws.cell(i, 11).number_format = _PCT
        for c in range(1, 12):
            ws.cell(i, c).font = Font(name=_FONT, size=10)
            ws.cell(i, c).border = _BORDER
            if i % 2 == 0:
                ws.cell(i, c).fill = _ZEBRA_FILL

    # Linha de totais
    tr = last + 1
    ws.cell(tr, 1, "TOTAL")
    for c, col in zip((6, 7, 8, 9, 10), "FGHIJ"):
        ws.cell(tr, c).value = f"=SUM({col}2:{col}{last})" if n else 0
        ws.cell(tr, c).number_format = _BRL
    ws.cell(tr, 11).value = f'=IFERROR(J{tr}/I{tr},"-")'
    ws.cell(tr, 11).number_format = _PCT
    for c in range(1, 12):
        cell = ws.cell(tr, c)
        cell.font = Font(name=_FONT, bold=True, size=10)
        cell.fill = _TOTAL_FILL
        cell.border = _BORDER

    for col, w in zip(range(1, 12), [11, 12, 20, 18, 16, 14, 14, 12, 14, 14, 11]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{tr}"

    # ---------------- Aba Resumo ----------------
    rs = wb.create_sheet("Resumo")
    D = f"Pedidos!$D$2:$D${last}"   # loja
    A = f"Pedidos!$A$2:$A${last}"   # data
    F = f"Pedidos!$F$2:$F${last}"; G = f"Pedidos!$G$2:$G${last}"
    H = f"Pedidos!$H$2:$H${last}"; I = f"Pedidos!$I$2:$I${last}"
    J = f"Pedidos!$J$2:$J${last}"

    periodo = f"{inicio.isoformat()} a {fim.isoformat()}"
    rs.cell(1, 1, f"Rentabilidade {periodo} — por Loja").font = Font(name=_FONT, bold=True, size=12)
    _header(rs, 2, ["Loja", "Pedidos", "Base (R$)", "Comissão (R$)",
                    "Frete (R$)", "Custo (R$)", "Lucro (R$)", "Margem %"])
    lojas = sorted({_loja_label(r["loja"], store_map) for r in rows})
    r0 = 3
    for idx, loja in enumerate(lojas):
        row = r0 + idx
        rs.cell(row, 1, loja)
        rs.cell(row, 2, f"=COUNTIF({D},$A{row})")
        rs.cell(row, 3, f"=SUMIF({D},$A{row},{F})")
        rs.cell(row, 4, f"=SUMIF({D},$A{row},{G})")
        rs.cell(row, 5, f"=SUMIF({D},$A{row},{H})")
        rs.cell(row, 6, f"=SUMIF({D},$A{row},{I})")
        rs.cell(row, 7, f"=SUMIF({D},$A{row},{J})")
        rs.cell(row, 8, f'=IFERROR(G{row}/F{row},"-")')
    lt = r0 + len(lojas)
    rs.cell(lt, 1, "TOTAL")
    rs.cell(lt, 2, f"=SUM(B{r0}:B{lt-1})" if lojas else 0)
    for col in "CDEFG":
        rs.cell(lt, ord(col) - 64, f"=SUM({col}{r0}:{col}{lt-1})" if lojas else 0)
    rs.cell(lt, 8, f'=IFERROR(G{lt}/F{lt},"-")')

    # por Dia (ao lado)
    b2 = 10
    rs.cell(1, b2, f"{periodo} — por Dia").font = Font(name=_FONT, bold=True, size=12)
    _header(rs, 2, ["Data", "Pedidos", "Lucro (R$)", "Margem %"], start=b2)
    dias = sorted({
        (r["data_sp"].isoformat() if hasattr(r["data_sp"], "isoformat") else str(r["data_sp"]))
        for r in rows
    })
    L = get_column_letter(b2)
    for idx, d in enumerate(dias):
        row = r0 + idx
        rs.cell(row, b2, d)
        rs.cell(row, b2 + 1, f"=COUNTIF({A},${L}{row})")
        rs.cell(row, b2 + 2, f"=SUMIF({A},${L}{row},{J})")
        rs.cell(row, b2 + 3, f'=IFERROR(SUMIF({A},${L}{row},{J})/SUMIF({A},${L}{row},{I}),"-")')
    dt = r0 + len(dias)
    rs.cell(dt, b2, "TOTAL")
    rs.cell(dt, b2 + 1, f"=SUM({get_column_letter(b2+1)}{r0}:{get_column_letter(b2+1)}{dt-1})" if dias else 0)
    rs.cell(dt, b2 + 2, f"=SUM({get_column_letter(b2+2)}{r0}:{get_column_letter(b2+2)}{dt-1})" if dias else 0)

    for row in range(3, lt + 1):
        for col in (3, 4, 5, 6, 7):
            rs.cell(row, col).number_format = _BRL
        rs.cell(row, 8).number_format = _PCT
    for row in range(3, dt + 1):
        rs.cell(row, b2 + 2).number_format = _BRL
        rs.cell(row, b2 + 3).number_format = _PCT
    for row in (lt, dt):
        for col in list(range(1, 9)) + list(range(b2, b2 + 4)):
            rs.cell(row, col).fill = _TOTAL_FILL
            rs.cell(row, col).font = Font(name=_FONT, bold=True, size=10)
    for row in range(3, max(lt, dt) + 1):
        for col in list(range(1, 9)) + list(range(b2, b2 + 4)):
            cell = rs.cell(row, col)
            if not (cell.font and cell.font.bold):
                cell.font = Font(name=_FONT, size=10)
            cell.border = _BORDER
    for col, w in zip(range(1, 9), [20, 9, 14, 14, 12, 14, 14, 11]):
        rs.column_dimensions[get_column_letter(col)].width = w
    for col, w in zip(range(b2, b2 + 4), [11, 9, 14, 11]):
        rs.column_dimensions[get_column_letter(col)].width = w
    rs.freeze_panes = "A3"

    # Força Excel/LibreOffice a recalcular as fórmulas ao abrir.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcId = 0
    except Exception:  # noqa: BLE001
        pass

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
