"""Excel de CONFERÊNCIA de um lote de importação avulsa (NfCommand).

Não é a planilha que a marionete sobe no destino — é uma visão RESUMIDA pra o
humano analisar/conferir o lote: uma linha por SKU (agregado), com modelo,
descrição, quantidade e valor total. Só entra o que tem informação (linhas com
quantidade/total > 0).

Camada de arquivo pura: recebe as linhas já agregadas (qualquer objeto com
`.modelo`/`.sku`/`.nome`/`.quantidade`/`.valor_total`) e devolve os bytes do
.xlsx. Não lê banco.
"""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Iterable, Protocol

from openpyxl import Workbook
from openpyxl.styles import Font

LOTE_EXCEL_MEDIA = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_SHEET = "lote"
_HEADERS = ("Modelo", "Bling SKU", "Descrição", "Quantidade", "Valor Total")


class _LinhaAgregada(Protocol):
    modelo: str | None
    sku: str
    nome: str
    quantidade: int
    valor_total: Decimal


def montar_xlsx(linhas: Iterable[_LinhaAgregada]) -> bytes:
    """Monta o .xlsx de conferência do lote (uma linha por SKU agregado)."""
    wb = Workbook()
    ws = wb.active
    ws.title = _SHEET

    ws.append(list(_HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)

    total_qtd = 0
    total_valor = Decimal("0.00")
    for ln in linhas:
        qtd = int(ln.quantidade or 0)
        valor = ln.valor_total if isinstance(ln.valor_total, Decimal) else Decimal(
            str(ln.valor_total or 0)
        )
        total_qtd += qtd
        total_valor += valor
        ws.append([
            ln.modelo or "",
            ln.sku or "",
            ln.nome or "",
            qtd,
            float(valor),
        ])

    # Linha de total no rodapé.
    total_row = ["", "", "Total", total_qtd, float(total_valor)]
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Largura amigável das colunas.
    for col, width in zip("ABCDE", (12, 16, 48, 12, 14)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
